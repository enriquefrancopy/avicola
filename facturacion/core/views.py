from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, F, Q
from django.http import JsonResponse
from django.db import models
from django.forms import modelformset_factory
from .models import Producto, Proveedor, Cliente, Factura, DetalleFactura, Pago, PagoFactura, Notificacion, MovimientoStock, ConfiguracionSistema, Caja, Gasto, MovimientoCaja, Denominacion
from .forms import ProductoForm, ProductoCrearForm, ProveedorForm, ClienteForm, FacturaForm, DetalleFacturaFormSet, PagoForm, PagoMultipleForm, AsignacionPagoForm, PagoFacturaFormSet

@login_required
def dashboard(request):
    """Vista del dashboard principal"""
    from datetime import datetime, timedelta
    from django.db.models import Count, Avg
    
    # Fechas para filtros
    hoy = datetime.now().date()
    inicio_mes = hoy.replace(day=1)
    inicio_anio = hoy.replace(month=1, day=1)
    
    # Métricas principales
    total_productos = Producto.objects.filter(activo=True).count()
    total_proveedores = Proveedor.objects.filter(activo=True).count()
    facturas_pendientes = Factura.objects.filter(estado='pendiente').count()
    total_por_pagar = Factura.objects.filter(estado='pendiente').aggregate(total=Sum('total'))['total'] or 0
    productos_stock_bajo_count = Producto.objects.filter(activo=True, stock__gt=0, stock__lte=F('stock_minimo')).count()
    productos_stock_bajo = Producto.objects.filter(activo=True, stock__gt=0, stock__lte=F('stock_minimo'))
    
    # Métricas adicionales
    facturas_mes = Factura.objects.filter(fecha__gte=inicio_mes).count()
    total_facturas_mes = Factura.objects.filter(fecha__gte=inicio_mes).aggregate(total=Sum('total'))['total'] or 0
    facturas_pagadas = Factura.objects.filter(estado='pagada').count()
    total_pagado = Factura.objects.filter(estado='pagada').aggregate(total=Sum('total'))['total'] or 0
    
    # Productos con stock crítico (stock = 0)
    productos_agotados_count = Producto.objects.filter(activo=True, stock=0).count()
    productos_agotados = Producto.objects.filter(activo=True, stock=0)
    
    # Top 5 productos con más stock
    productos_top_stock = Producto.objects.filter(activo=True).order_by('-stock')[:5]
    
    # Top 5 proveedores por saldo
    proveedores_top_saldo = Proveedor.objects.filter(activo=True).order_by('-saldo')[:5]
    
    # Facturas recientes
    facturas_recientes = Factura.objects.select_related('proveedor').order_by('-fecha')[:5]
    
    # Datos para gráficos
    # Facturas por mes (últimos 12 meses)
    meses_grafico = []
    facturas_por_mes = []
    montos_por_mes = []
    
    for i in range(12):
        fecha = hoy - timedelta(days=30*i)
        inicio_mes_grafico = fecha.replace(day=1)
        fin_mes_grafico = (inicio_mes_grafico + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Solo facturas de venta (no compras)
        facturas_mes_count = Factura.objects.filter(
            fecha__gte=inicio_mes_grafico, 
            fecha__lte=fin_mes_grafico,
            tipo='venta'
        ).count()
        
        facturas_mes_total = Factura.objects.filter(
            fecha__gte=inicio_mes_grafico, 
            fecha__lte=fin_mes_grafico,
            tipo='venta'
        ).aggregate(total=Sum('total'))['total'] or 0
        
        meses_grafico.insert(0, fecha.strftime('%b %Y'))
        facturas_por_mes.insert(0, facturas_mes_count)
        montos_por_mes.insert(0, float(facturas_mes_total))
    
    # Distribución de estados de facturas
    estados_facturas = Factura.objects.values('estado').annotate(count=Count('estado'))
    estados_labels = [item['estado'].title() for item in estados_facturas]
    estados_data = [item['count'] for item in estados_facturas]
    
    # Productos por categoría de stock
    productos_stock_alto_count = Producto.objects.filter(activo=True, stock__gt=F('stock_minimo')).count()
    productos_stock_alto = Producto.objects.filter(activo=True, stock__gt=F('stock_minimo'))
    
    # Obtener alertas del sistema
    alertas = obtener_alertas_stock(request)
    
    context = {
        # Fecha actual
        'fecha_actual': datetime.now(),
        
        # Métricas principales
        'total_productos': total_productos,
        'total_proveedores': total_proveedores,
        'facturas_pendientes': facturas_pendientes,
        'total_por_pagar': total_por_pagar,
        'productos_stock_bajo': productos_stock_bajo,
        'productos_stock_bajo_count': productos_stock_bajo_count,
        
        # Métricas adicionales
        'facturas_mes': facturas_mes,
        'total_facturas_mes': total_facturas_mes,
        'facturas_pagadas': facturas_pagadas,
        'total_pagado': total_pagado,
        'productos_agotados': productos_agotados,
        'productos_agotados_count': productos_agotados_count,
        
        # Listas
        'productos_top_stock': productos_top_stock,
        'proveedores_top_saldo': proveedores_top_saldo,
        'facturas_recientes': facturas_recientes,
        'notificaciones': Notificacion.objects.filter(leida=False)[:5],
        
        # Alertas del sistema
        'alertas': alertas,
        'total_alertas': alertas['total_alertas'],
        
        # Datos para gráficos
        'meses_grafico': meses_grafico,
        'facturas_por_mes': facturas_por_mes,
        'montos_por_mes': montos_por_mes,
        'estados_labels': estados_labels,
        'estados_data': estados_data,
        'productos_stock_alto': productos_stock_alto,
        'productos_stock_alto_count': productos_stock_alto_count,
        
        # Productos por categoría de stock para gráfico
        'productos_stock_medio_count': Producto.objects.filter(
            activo=True, 
            stock__gt=F('stock_minimo'), 
            stock__lte=F('stock_minimo') * 2
        ).count(),
    }
    
    return render(request, 'dashboard.html', context)

@login_required
def dashboard_data(request):
    """Vista AJAX para obtener datos del dashboard"""
    from datetime import datetime, timedelta
    from django.http import JsonResponse
    
    # Obtener parámetros
    mes_seleccionado = request.GET.get('mes')
    anio_seleccionado = request.GET.get('anio')
    
    if mes_seleccionado and anio_seleccionado:
        # Datos para un mes específico
        try:
            fecha_inicio = datetime(int(anio_seleccionado), int(mes_seleccionado), 1).date()
            fecha_fin = (fecha_inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            # Ventas diarias del mes
            ventas_diarias = []
            for dia in range(1, fecha_fin.day + 1):
                fecha_dia = fecha_inicio.replace(day=dia)
                fecha_dia_inicio = datetime.combine(fecha_dia, datetime.min.time())
                fecha_dia_fin = datetime.combine(fecha_dia, datetime.max.time())
                
                total_dia = Factura.objects.filter(
                    fecha__gte=fecha_dia_inicio,
                    fecha__lte=fecha_dia_fin,
                    tipo='venta'
                ).aggregate(total=Sum('total'))['total'] or 0
                
                ventas_diarias.append({
                    'dia': dia,
                    'fecha': fecha_dia.strftime('%d/%m/%Y'),
                    'total': float(total_dia)
                })
            
            return JsonResponse({
                'success': True,
                'ventas_diarias': ventas_diarias,
                'mes_nombre': fecha_inicio.strftime('%B %Y')
            })
            
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Fecha inválida'})
    
    else:
        # Datos de facturación mensual (últimos 12 meses)
        hoy = datetime.now().date()
        meses_data = []
        
        for i in range(12):
            fecha = hoy - timedelta(days=30*i)
            inicio_mes = fecha.replace(day=1)
            fin_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            total_mes = Factura.objects.filter(
                fecha__gte=inicio_mes,
                fecha__lte=fin_mes,
                tipo='venta'
            ).aggregate(total=Sum('total'))['total'] or 0
            
            meses_data.append({
                'mes': fecha.strftime('%b %Y'),
                'total': float(total_mes),
                'anio': fecha.year,
                'mes_num': fecha.month
            })
        
        return JsonResponse({
            'success': True,
            'facturacion_mensual': meses_data
        })

@login_required
def productos_list(request):
    """Lista de productos"""
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    
    # Base query - mostrar todos los productos
    productos = Producto.objects.all().order_by('nombre')
    
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | Q(codigo__icontains=q)
        )
    
    if estado:
        if estado == 'normal':
            productos = productos.filter(stock__gt=F('stock_minimo'))
        elif estado == 'minimo':
            productos = productos.filter(stock=F('stock_minimo'))
        elif estado == 'critico':
            productos = productos.filter(stock__lt=F('stock_minimo'))
    
    return render(request, 'productos_list.html', {
        'productos': productos,
        'productos_stock_bajo': Producto.objects.filter(activo=True, stock__lte=F('stock_minimo')).count()
    })

@login_required
def producto_crear(request):
    """Crear un nuevo producto"""
    if request.method == 'POST':
        form = ProductoCrearForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, 'Producto creado correctamente.')
            return redirect('productos_list')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    return redirect('productos_list')

@login_required
def producto_editar(request, pk):
    """Editar un producto existente"""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Producto actualizado correctamente.')
                return redirect('productos_list')
            except Exception as e:
                messages.error(request, f'Error al guardar el producto: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    return redirect('productos_list')

@login_required
def producto_eliminar(request, pk):
    """Desactivar un producto"""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.activo = False
        producto.save()
        messages.success(request, f'Producto "{producto.nombre}" desactivado correctamente.')
    return redirect('productos_list')

@login_required
def producto_reactivar(request, pk):
    """Reactivar un producto"""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.activo = True
        producto.save()
        messages.success(request, f'Producto "{producto.nombre}" reactivado correctamente.')
    return redirect('productos_list')

@login_required
def producto_ajustar_stock(request, pk):
    """Ajustar el stock de un producto"""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        tipo_movimiento = request.POST.get('tipo_movimiento')
        observacion = request.POST.get('observacion', '')
        try:
            cantidad = int(request.POST.get('cantidad', 0))
            if cantidad <= 0:
                raise ValueError('La cantidad debe ser mayor a 0')
            
            # Registrar el movimiento de stock
            if tipo_movimiento == 'entrada':
                MovimientoStock.registrar_movimiento(
                    producto=producto,
                    tipo='entrada',
                    origen='ajuste_manual',
                    cantidad=cantidad,
                    usuario=request.user,
                    observacion=observacion
                )
            elif tipo_movimiento == 'salida':
                if producto.stock < cantidad:
                    raise ValueError('No hay suficiente stock disponible')
                MovimientoStock.registrar_movimiento(
                    producto=producto,
                    tipo='salida',
                    origen='ajuste_manual',
                    cantidad=cantidad,
                    usuario=request.user,
                    observacion=observacion
                )
            
            # Crear notificación si el stock es bajo
            if producto.stock <= producto.stock_minimo:
                Notificacion.objects.create(
                    mensaje=f'Stock bajo en producto {producto.nombre} ({producto.stock} unidades)',
                    tipo='warning'
                )
            
            messages.success(request, f'Stock ajustado correctamente. Nuevo stock: {producto.stock}')
        except ValueError as e:
            messages.error(request, str(e))
        
    return redirect('productos_list')

@login_required
def recalcular_totales_factura(request, pk):
    """Recalcular los totales de una factura existente"""
    factura = get_object_or_404(Factura, pk=pk)
    
    # Recalcular totales
    detalles = factura.detalles.all()
    
    if detalles.count() > 0:
        # Calcular subtotales discriminados por IVA
        subtotal_5 = sum(d.subtotal for d in detalles if d.producto.iva == 5)
        subtotal_10 = sum(d.subtotal for d in detalles if d.producto.iva == 10)
        subtotal_total = subtotal_5 + subtotal_10
        
        iva = sum(d.iva for d in detalles)
        total = subtotal_total  # Total es la suma de subtotales (sin IVA)
        
        print(f"Recalculando factura {pk}: subtotal_5={subtotal_5}, subtotal_10={subtotal_10}, subtotal_total={subtotal_total}, iva={iva}, total={total}")
        
        factura.subtotal = subtotal_total
        factura.iva = iva
        factura.total = total
        factura.save()
        
        messages.success(request, f'Totales de la factura #{pk} recalculados correctamente.')
    else:
        messages.warning(request, f'La factura #{pk} no tiene detalles para recalcular.')
    
    return redirect('factura_ver', pk=pk)



@login_required
def proveedores_list(request):
    """Lista de proveedores"""
    proveedores = Proveedor.objects.all()
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    
    if q:
        proveedores = proveedores.filter(
            Q(nombre__icontains=q) | Q(rif__icontains=q) | Q(email__icontains=q)
        )
    
    if estado:
        if estado == 'activo':
            proveedores = proveedores.filter(activo=True)
        elif estado == 'inactivo':
            proveedores = proveedores.filter(activo=False)
    
    return render(request, 'proveedores_list.html', {'proveedores': proveedores})

@login_required
def proveedor_crear(request):
    """Crear un nuevo proveedor"""
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, 'Proveedor creado correctamente.')
            return redirect('proveedores_list')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    return redirect('proveedores_list')

@login_required
def proveedor_editar(request, pk):
    """Editar un proveedor existente"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado correctamente.')
            return redirect('proveedores_list')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    return redirect('proveedores_list')

@login_required
def proveedor_eliminar(request, pk):
    """Desactivar un proveedor"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.activo = False
        proveedor.save()
        messages.success(request, f'Proveedor "{proveedor.nombre}" desactivado correctamente.')
        return redirect('proveedores_list')
    return redirect('proveedores_list')

@login_required
def proveedor_reactivar(request, pk):
    """Reactivar un proveedor"""
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.activo = True
        proveedor.save()
        messages.success(request, f'Proveedor "{proveedor.nombre}" reactivado correctamente.')
        return redirect('proveedores_list')
    return redirect('proveedores_list')

@login_required
def factura_list(request):
    """Lista de facturas"""
    tipo = request.GET.get('tipo', 'compra')  # Por defecto mostrar compras
    facturas = Factura.objects.filter(tipo=tipo)
    
    # Filtros
    q = request.GET.get('q', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        if tipo == 'compra':
            facturas = facturas.filter(
                Q(proveedor__nombre__icontains=q) | Q(numero__icontains=q)
            )
        else:
            facturas = facturas.filter(
                Q(cliente__nombre__icontains=q) | Q(numero__icontains=q)
            )
    
    if desde:
        facturas = facturas.filter(fecha__gte=desde)
    
    if hasta:
        facturas = facturas.filter(fecha__lte=hasta)
    
    facturas = facturas.order_by('-fecha')
    
    return render(request, 'factura_list.html', {
        'facturas': facturas,
        'tipo_actual': tipo,
        'tipos': [('compra', 'Compras'), ('venta', 'Ventas')]
    })

@login_required
def factura_crear(request):
    """Crear una nueva factura"""
    from datetime import datetime
    
    tipo = request.GET.get('tipo', 'compra')
    fecha_actual = datetime.now().strftime('%Y-%m-%dT%H:%M')
    
    if request.method == 'POST':
        # Crear factura manualmente
        tipo = request.POST.get('tipo')
        fecha = request.POST.get('fecha')
        proveedor_id = request.POST.get('proveedor')
        cliente_id = request.POST.get('cliente')
        observacion = request.POST.get('observacion')
        
        # Validar campos requeridos
        if tipo == 'compra' and not proveedor_id:
            messages.error(request, 'Debe seleccionar un proveedor para facturas de compra.')
            return render(request, 'factura_form.html', {
                'tipo': tipo,
                'titulo': f'Nueva Factura de {tipo.title()}',
                'fecha_actual': fecha_actual
            })
        
        if tipo == 'venta' and not cliente_id:
            messages.error(request, 'Debe seleccionar un cliente para facturas de venta.')
            return render(request, 'factura_form.html', {
                'tipo': tipo,
                'titulo': f'Nueva Factura de {tipo.title()}',
                'fecha_actual': fecha_actual
            })
        
        # Crear factura
        factura = Factura()
        factura.tipo = tipo
        factura.fecha = fecha
        factura.observacion = observacion
        factura.usuario = request.user
        
        if tipo == 'compra':
            factura.proveedor_id = proveedor_id
        else:
            factura.cliente_id = cliente_id
        
        # Inicializar totales en 0
        factura.subtotal = 0
        factura.iva = 0
        factura.total = 0
        
        factura.save()
        
        # Procesar detalles
        detalles_data = []
        i = 0
        print(f"POST data: {request.POST}")
        
        while f'detalles-{i}-producto' in request.POST:
            producto_id = request.POST.get(f'detalles-{i}-producto')
            cantidad = request.POST.get(f'detalles-{i}-cantidad')
            precio_unitario = request.POST.get(f'detalles-{i}-precio_unitario')
            precio_venta = request.POST.get(f'detalles-{i}-precio_venta')  # Nuevo campo para precio de venta
            
            print(f"Procesando detalle {i}: producto_id={producto_id}, cantidad={cantidad}, precio={precio_unitario}, precio_venta={precio_venta}")
            
            if producto_id and cantidad and precio_unitario:
                detalle = DetalleFactura()
                detalle.factura = factura
                detalle.producto_id = producto_id
                detalle.cantidad = cantidad
                detalle.precio_unitario = precio_unitario
                detalle.subtotal = int(cantidad) * int(precio_unitario)
                
                # Obtener el IVA del producto
                producto = Producto.objects.get(id=producto_id)
                if producto.iva == 5:
                    detalle.iva = int(detalle.subtotal / 21)
                else:  # IVA 10%
                    detalle.iva = int(detalle.subtotal / 11)
                
                # Calcular el total del detalle (subtotal + iva)
                detalle.total = detalle.subtotal + detalle.iva
                
                detalle.save()
                print(f"Detalle guardado: subtotal={detalle.subtotal}, iva={detalle.iva}")
                
                # Actualizar stock y registrar movimiento
                if tipo == 'compra':
                    # Registrar movimiento de entrada para compras
                    MovimientoStock.registrar_movimiento(
                        producto=producto,
                        tipo='entrada',
                        origen='factura_compra',
                        cantidad=int(cantidad),
                        usuario=request.user,
                        referencia=f'Factura #{factura.id}',
                        observacion=f'Compra de {cantidad} unidades a Gs. {precio_unitario} c/u'
                    )
                    
                    # Para facturas de compra, actualizar costo si es diferente
                    nuevo_costo = int(precio_unitario)
                    if producto.costo != nuevo_costo:
                        print(f"Actualizando costo del producto {producto.nombre}: {producto.costo} -> {nuevo_costo}")
                        producto.costo = nuevo_costo
                    
                    # Actualizar precio de venta si se proporcionó
                    if precio_venta:
                        nuevo_precio_venta = int(precio_venta)
                        if producto.precio != nuevo_precio_venta:
                            print(f"Actualizando precio de venta del producto {producto.nombre}: {producto.precio} -> {nuevo_precio_venta}")
                            producto.precio = nuevo_precio_venta
                    
                    producto.save()
                else:
                    # Registrar movimiento de salida para ventas
                    MovimientoStock.registrar_movimiento(
                        producto=producto,
                        tipo='salida',
                        origen='factura_venta',
                        cantidad=int(cantidad),
                        usuario=request.user,
                        referencia=f'Factura #{factura.id}',
                        observacion=f'Venta de {cantidad} unidades a Gs. {precio_unitario} c/u'
                    )
            else:
                print(f"Detalle {i} omitido: datos incompletos")
            
            i += 1
        
        print(f"Total de detalles procesados: {i}")
        
        # Verificar que la factura tenga al menos un producto
        if i == 0:
            # Eliminar la factura vacía
            factura.delete()
            messages.error(request, 'La factura debe tener al menos un producto en el detalle.')
            return render(request, 'factura_form.html', {
                'tipo': tipo,
                'titulo': f'Nueva Factura de {tipo.title()}',
                'fecha_actual': fecha_actual
            })
        
        # Calcular totales con IVA discriminado
        detalles = factura.detalles.all()
        print(f"Detalles encontrados: {detalles.count()}")
        
        if detalles.count() > 0:
            # Calcular subtotales discriminados por IVA
            subtotal_5 = sum(d.subtotal for d in detalles if d.producto.iva == 5)
            subtotal_10 = sum(d.subtotal for d in detalles if d.producto.iva == 10)
            subtotal_total = subtotal_5 + subtotal_10
            
            iva = sum(d.iva for d in detalles)
            total = subtotal_total  # Total es la suma de subtotales (sin IVA)
            
            print(f"Totales calculados: subtotal_5={subtotal_5}, subtotal_10={subtotal_10}, subtotal_total={subtotal_total}, iva={iva}, total={total}")
            
            factura.subtotal = subtotal_total
            factura.iva = iva
            factura.total = total
        else:
            print("No se encontraron detalles, manteniendo totales en 0")
        
        factura.save()
        
        # Actualizar saldo del proveedor o cliente según el tipo de factura
        if factura.tipo == 'compra' and factura.proveedor:
            factura.proveedor.saldo += factura.total
            factura.proveedor.save()
        elif factura.tipo == 'venta' and factura.cliente:
            factura.cliente.saldo += factura.total
            factura.cliente.save()
        
        messages.success(request, f'Factura de {factura.get_tipo_display()} creada correctamente.')
        return redirect(f'{reverse("factura_list")}?tipo={factura.tipo}')
    
    return render(request, 'factura_form.html', {
        'tipo': tipo,
        'titulo': f'Nueva Factura de {tipo.title()}',
        'fecha_actual': fecha_actual
    })

@login_required
def factura_editar(request, pk):
    """Editar una factura existente"""
    factura = get_object_or_404(Factura, pk=pk)
    if factura.estado != 'pendiente':
        messages.error(request, 'Solo se pueden editar facturas pendientes.')
        return redirect('factura_ver', pk=pk)
    
    if request.method == 'POST':
        form = FacturaForm(request.POST, instance=factura)
        if form.is_valid():
            factura = form.save()
            formset = DetalleFacturaFormSet(request.POST, instance=factura)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Factura actualizada correctamente.')
                return redirect('factura_ver', pk=pk)
    else:
        form = FacturaForm(instance=factura)
        formset = DetalleFacturaFormSet(instance=factura)
    
    return redirect('factura_list')

@login_required
def factura_eliminar(request, pk):
    """Eliminar una factura y descontar productos del stock"""
    factura = get_object_or_404(Factura, pk=pk)
    if request.method == 'POST':
        try:
            # Revertir movimientos de stock según el tipo de factura
            for detalle in factura.detalles.all():
                producto = detalle.producto
                if factura.tipo == 'compra':
                    # Si es factura de compra, registrar salida (reversión de entrada)
                    MovimientoStock.registrar_movimiento(
                        producto=producto,
                        tipo='salida',
                        origen='factura_compra',
                        cantidad=detalle.cantidad,
                        usuario=request.user,
                        referencia=f'Factura #{factura.numero} eliminada',
                        observacion='Eliminación de factura de compra'
                    )
                else:
                    # Si es factura de venta, registrar entrada (reversión de salida)
                    MovimientoStock.registrar_movimiento(
                        producto=producto,
                        tipo='entrada',
                        origen='factura_venta',
                        cantidad=detalle.cantidad,
                        usuario=request.user,
                        referencia=f'Factura #{factura.numero} eliminada',
                        observacion='Eliminación de factura de venta'
                    )
            
            # Restaurar saldos de proveedor/cliente si hay pagos
            if factura.tipo == 'compra' and factura.proveedor:
                factura.proveedor.saldo -= factura.total_pagado
                factura.proveedor.save()
            elif factura.tipo == 'venta' and factura.cliente:
                factura.cliente.saldo -= factura.total_pagado
                factura.cliente.save()
            
            # Eliminar la factura (esto también eliminará los pagos por CASCADE)
            factura.delete()
            
            messages.success(request, 'Factura eliminada correctamente.')
            return redirect('factura_list')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar la factura: {str(e)}')
            return redirect('factura_ver', pk=pk)
    
    return render(request, 'factura_confirm_eliminar.html', {'factura': factura})

@login_required
def factura_anular(request, pk):
    """Anular una factura cambiando su estado a 'anulada'"""
    factura = get_object_or_404(Factura, pk=pk)
    
    # Verificar que la factura no esté ya anulada
    if factura.estado == 'anulada':
        messages.warning(request, 'La factura ya está anulada.')
        return redirect('factura_ver', pk=pk)
    
    if request.method == 'POST':
        try:
            # Revertir movimientos de stock antes de anular
            for detalle in factura.detalles.all():
                if factura.tipo == 'compra':
                    # Para compras anuladas, registrar salida (reversión de entrada)
                    MovimientoStock.registrar_movimiento(
                        producto=detalle.producto,
                        tipo='salida',
                        origen='factura_compra',
                        cantidad=detalle.cantidad,
                        usuario=request.user,
                        referencia=f'Factura #{factura.id} (ANULADA)',
                        observacion=f'Anulación: Reversión de compra de {detalle.cantidad} unidades'
                    )
                else:
                    # Para ventas anuladas, registrar entrada (reversión de salida)
                    MovimientoStock.registrar_movimiento(
                        producto=detalle.producto,
                        tipo='entrada',
                        origen='factura_venta',
                        cantidad=detalle.cantidad,
                        usuario=request.user,
                        referencia=f'Factura #{factura.id} (ANULADA)',
                        observacion=f'Anulación: Reversión de venta de {detalle.cantidad} unidades'
                    )
            
            # Actualizar saldo del proveedor o cliente antes de anular
            if factura.tipo == 'compra' and factura.proveedor and factura.estado == 'pendiente':
                # Reducir el saldo del proveedor ya que la factura ya no está pendiente
                factura.proveedor.saldo -= factura.total
                factura.proveedor.save()
            elif factura.tipo == 'venta' and factura.cliente and factura.estado == 'pendiente':
                # Reducir el saldo del cliente ya que la factura ya no está pendiente
                factura.cliente.saldo -= factura.total
                factura.cliente.save()
            
            # Cambiar el estado a anulada
            factura.estado = 'anulada'
            factura.save(update_fields=['estado'])
            
            messages.success(request, 'Factura anulada correctamente.')
            return redirect('factura_ver', pk=pk)
            
        except Exception as e:
            messages.error(request, f'Error al anular la factura: {str(e)}')
            return redirect('factura_ver', pk=pk)
    
    return render(request, 'factura_confirm_anular.html', {'factura': factura})

@login_required
def factura_ver(request, pk):
    """Ver detalle de factura"""
    factura = get_object_or_404(Factura, pk=pk)
    return render(request, 'factura_ver.html', {'factura': factura})

@login_required
def factura_pagos(request, pk):
    """Ver y gestionar pagos de una factura"""
    factura = get_object_or_404(Factura, pk=pk)
    pagos_facturas = factura.pagos_facturas.select_related('pago', 'pago__usuario').all()
    
    # Obtener caja activa del día
    caja_activa = Caja.obtener_caja_activa()
    
    if request.method == 'POST':
        form = PagoForm(request.POST, factura=factura)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.factura = factura
            pago.usuario = request.user
            pago.save()
            
            # Actualizar saldo del proveedor o cliente según el tipo de factura
            if factura.tipo == 'compra' and factura.proveedor:
                factura.proveedor.saldo -= pago.monto_total
                factura.proveedor.save()
            elif factura.tipo == 'venta' and factura.cliente:
                factura.cliente.saldo -= pago.monto_total
                factura.cliente.save()
            
            # Actualizar estado de la factura
            factura.actualizar_estado()
            
            # Mensaje informativo con detalles del pago
            monto_billete = form.cleaned_data.get('monto_billete')
            vuelto = form.cleaned_data.get('vuelto', 0)
            
            if factura.tipo == 'venta' and monto_billete:
                mensaje = f'Pago registrado correctamente. Monto recibido: Gs. {monto_billete:,}, Vuelto: Gs. {vuelto:,}'
            else:
                mensaje = 'Pago registrado correctamente.'
            
            # Agregar información sobre la caja
            if caja_activa:
                if factura.tipo == 'venta':
                    mensaje += f' Se ha creado automáticamente un ingreso en la caja del día.'
                else:
                    mensaje += f' Se ha creado automáticamente un egreso en la caja del día.'
            
            messages.success(request, mensaje)
            
            return redirect('factura_pagos', pk=pk)
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
            # También mostrar errores de campos específicos
            for field_name, field in form.fields.items():
                if field_name in form.errors:
                    for error in form.errors[field_name]:
                        messages.error(request, f'Error en {field_name}: {error}')
    else:
        form = PagoForm(factura=factura)
    
    return render(request, 'factura_pagos.html', {
        'factura': factura,
        'pagos_facturas': pagos_facturas,
        'form': form,
        'caja_activa': caja_activa
    })

@login_required
def stock_movimientos(request):
    """Vista de movimientos de stock"""
    movimientos = MovimientoStock.objects.select_related('producto', 'usuario').all().order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')
    origen = request.GET.get('origen', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        movimientos = movimientos.filter(
            Q(producto__nombre__icontains=q) | Q(producto__codigo__icontains=q)
        )
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    if origen:
        movimientos = movimientos.filter(origen=origen)
    
    if desde:
        movimientos = movimientos.filter(fecha__date__gte=desde)
    
    if hasta:
        movimientos = movimientos.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(movimientos, 50)  # 50 movimientos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'movimientos': page_obj,
        'tipos_movimiento': MovimientoStock.TIPO_CHOICES,
        'origenes_movimiento': MovimientoStock.ORIGEN_CHOICES,
    }
    
    return render(request, 'stock_movimientos.html', context)

# ============================================================================
# PAGOS
# ============================================================================

@login_required
def pagos_dashboard(request):
    """Dashboard principal del módulo de pagos"""
    # Obtener estadísticas de pagos
    total_pagos_clientes = Pago.objects.filter(factura__tipo='venta').count()
    total_pagos_proveedores = Pago.objects.filter(factura__tipo='compra').count()
    
    # Pagos recientes
    pagos_recientes = Pago.objects.select_related('factura', 'usuario').order_by('-fecha')[:10]
    
    # Facturas pendientes
    facturas_pendientes_clientes = Factura.objects.filter(
        tipo='venta', 
        estado='pendiente'
    ).select_related('cliente').order_by('-fecha')[:5]
    
    facturas_pendientes_proveedores = Factura.objects.filter(
        tipo='compra', 
        estado='pendiente'
    ).select_related('proveedor').order_by('-fecha')[:5]
    
    # Caja activa
    caja_activa = Caja.obtener_caja_activa()
    
    context = {
        'total_pagos_clientes': total_pagos_clientes,
        'total_pagos_proveedores': total_pagos_proveedores,
        'pagos_recientes': pagos_recientes,
        'facturas_pendientes_clientes': facturas_pendientes_clientes,
        'facturas_pendientes_proveedores': facturas_pendientes_proveedores,
        'caja_activa': caja_activa,
    }
    
    return render(request, 'pagos_dashboard.html', context)

@login_required
def pagos_clientes_list(request):
    """Lista de pagos de clientes (facturas de venta)"""
    # Obtener facturas de venta con pagos
    facturas_venta = Factura.objects.filter(
        tipo='venta'
    ).select_related('cliente').prefetch_related('pagos').order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        facturas_venta = facturas_venta.filter(
            Q(cliente__nombre__icontains=q) | 
            Q(numero__icontains=q) |
            Q(cliente__ruc__icontains=q)
        )
    
    if estado:
        facturas_venta = facturas_venta.filter(estado=estado)
    
    if desde:
        facturas_venta = facturas_venta.filter(fecha__date__gte=desde)
    
    if hasta:
        facturas_venta = facturas_venta.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(facturas_venta, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'facturas': page_obj,
        'tipo': 'venta',
        'titulo': 'Pagos de Clientes',
        'subtitulo': 'Facturas de Venta - Gestión de Cobros',
    }
    
    return render(request, 'pagos_list.html', context)

@login_required
def pagos_proveedores_list(request):
    """Lista de pagos a proveedores (facturas de compra)"""
    # Obtener facturas de compra con pagos
    facturas_compra = Factura.objects.filter(
        tipo='compra'
    ).select_related('proveedor').prefetch_related('pagos').order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        facturas_compra = facturas_compra.filter(
            Q(proveedor__nombre__icontains=q) | 
            Q(numero__icontains=q) |
            Q(proveedor__ruc__icontains=q)
        )
    
    if estado:
        facturas_compra = facturas_compra.filter(estado=estado)
    
    if desde:
        facturas_compra = facturas_compra.filter(fecha__date__gte=desde)
    
    if hasta:
        facturas_compra = facturas_compra.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(facturas_compra, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'facturas': page_obj,
        'tipo': 'compra',
        'titulo': 'Pagos a Proveedores',
        'subtitulo': 'Facturas de Compra - Gestión de Pagos',
    }
    
    return render(request, 'pagos_list.html', context)

# ============================================================================
# PAGOS
# ============================================================================

@login_required
def pagos_dashboard(request):
    """Dashboard principal del módulo de pagos"""
    # Obtener estadísticas de pagos
    total_pagos_clientes = Pago.objects.filter(factura__tipo='venta').count()
    total_pagos_proveedores = Pago.objects.filter(factura__tipo='compra').count()
    
    # Pagos recientes
    pagos_recientes = Pago.objects.select_related('factura', 'usuario').order_by('-fecha')[:10]
    
    # Facturas pendientes
    facturas_pendientes_clientes = Factura.objects.filter(
        tipo='venta', 
        estado='pendiente'
    ).select_related('cliente').order_by('-fecha')[:5]
    
    facturas_pendientes_proveedores = Factura.objects.filter(
        tipo='compra', 
        estado='pendiente'
    ).select_related('proveedor').order_by('-fecha')[:5]
    
    # Caja activa
    caja_activa = Caja.obtener_caja_activa()
    
    context = {
        'total_pagos_clientes': total_pagos_clientes,
        'total_pagos_proveedores': total_pagos_proveedores,
        'pagos_recientes': pagos_recientes,
        'facturas_pendientes_clientes': facturas_pendientes_clientes,
        'facturas_pendientes_proveedores': facturas_pendientes_proveedores,
        'caja_activa': caja_activa,
    }
    
    return render(request, 'pagos_dashboard.html', context)

@login_required
def pagos_clientes_list(request):
    """Lista de pagos de clientes (facturas de venta)"""
    # Obtener facturas de venta con pagos
    facturas_venta = Factura.objects.filter(
        tipo='venta'
    ).select_related('cliente').prefetch_related('pagos').order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        facturas_venta = facturas_venta.filter(
            Q(cliente__nombre__icontains=q) | 
            Q(numero__icontains=q) |
            Q(cliente__ruc__icontains=q)
        )
    
    if estado:
        facturas_venta = facturas_venta.filter(estado=estado)
    
    if desde:
        facturas_venta = facturas_venta.filter(fecha__date__gte=desde)
    
    if hasta:
        facturas_venta = facturas_venta.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(facturas_venta, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'facturas': page_obj,
        'tipo': 'venta',
        'titulo': 'Pagos de Clientes',
        'subtitulo': 'Facturas de Venta - Gestión de Cobros',
    }
    
    return render(request, 'pagos_list.html', context)

@login_required
def pagos_proveedores_list(request):
    """Lista de pagos a proveedores (facturas de compra)"""
    # Obtener facturas de compra con pagos
    facturas_compra = Factura.objects.filter(
        tipo='compra'
    ).select_related('proveedor').prefetch_related('pagos').order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        facturas_compra = facturas_compra.filter(
            Q(proveedor__nombre__icontains=q) | 
            Q(numero__icontains=q) |
            Q(proveedor__ruc__icontains=q)
        )
    
    if estado:
        facturas_compra = facturas_compra.filter(estado=estado)
    
    if desde:
        facturas_compra = facturas_compra.filter(fecha__date__gte=desde)
    
    if hasta:
        facturas_compra = facturas_compra.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(facturas_compra, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'facturas': page_obj,
        'tipo': 'compra',
        'titulo': 'Pagos a Proveedores',
        'subtitulo': 'Facturas de Compra - Gestión de Pagos',
    }
    
    return render(request, 'pagos_list.html', context)

# Pagos
@login_required
def pago_crear(request, factura_id):
    """Crear un nuevo pago"""
    factura = get_object_or_404(Factura, pk=factura_id)
    
    # Obtener caja activa del día
    caja_activa = Caja.obtener_caja_activa()
    
    if request.method == 'POST':
        form = PagoForm(request.POST, factura=factura)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.factura = factura
            pago.usuario = request.user
            pago.save()
            
            # Actualizar saldo del proveedor o cliente según el tipo de factura
            if factura.tipo == 'compra' and factura.proveedor:
                factura.proveedor.saldo -= pago.monto_total
                factura.proveedor.save()
            elif factura.tipo == 'venta' and factura.cliente:
                factura.cliente.saldo -= pago.monto_total
                factura.cliente.save()
            
            # Actualizar estado de la factura
            factura.actualizar_estado()
            
            # Mensaje informativo sobre la caja
            if caja_activa:
                if factura.tipo == 'venta':
                    messages.success(request, f'Pago registrado correctamente. Se ha creado automáticamente un ingreso en la caja del día.')
                else:
                    messages.success(request, f'Pago registrado correctamente. Se ha creado automáticamente un egreso en la caja del día.')
            else:
                messages.success(request, 'Pago registrado correctamente.')
            
            return redirect('factura_pagos', pk=factura_id)
    else:
        form = PagoForm(factura=factura)
    
    return render(request, 'factura_pagos.html', {
        'factura': factura,
        'pagos': factura.pagos.all(),
        'form': form,
        'caja_activa': caja_activa
    })

@login_required
def pago_eliminar(request, pk):
    """Eliminar un pago"""
    pago = get_object_or_404(Pago, pk=pk)
    factura_id = pago.factura.id
    if request.method == 'POST':
        # Restaurar saldo del proveedor o cliente antes de eliminar el pago
        if pago.factura.tipo == 'compra' and pago.factura.proveedor:
            pago.factura.proveedor.saldo += pago.monto
            pago.factura.proveedor.save()
        elif pago.factura.tipo == 'venta' and pago.factura.cliente:
            pago.factura.cliente.saldo += pago.monto
            pago.factura.cliente.save()
        
        pago.delete()
        
        # Actualizar estado de la factura después de eliminar el pago
        pago.factura.actualizar_estado()
        
        messages.success(request, 'Pago eliminado correctamente.')
        return redirect('factura_pagos', pk=factura_id)
    return render(request, 'pago_confirm_delete.html', {'pago': pago})

# API endpoints
@login_required
def get_producto_info(request):
    """Obtener información de un producto via AJAX"""
    producto_id = request.GET.get('id')
    try:
        producto = Producto.objects.get(id=producto_id)
        return JsonResponse({
            'precio': float(producto.precio),
            'iva': float(producto.iva),
            'stock': producto.stock
        })
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

@login_required
def calcular_total_factura(request):
    """Calcular totales de factura via AJAX"""
    subtotal = float(request.GET.get('subtotal', 0))
    iva = float(request.GET.get('iva', 0))
    total = subtotal + (subtotal * (iva / 100))
    return JsonResponse({'total': total})

@login_required
def buscar_proveedores(request):
    """Buscar proveedores para autocompletado"""
    q = request.GET.get('q', '').strip()
    print(f"Búsqueda de proveedores: '{q}'")
    
    try:
        if len(q) >= 2:
            # Buscar por nombre o RIF (ignorando mayúsculas/minúsculas)
            proveedores = Proveedor.objects.filter(
                Q(nombre__icontains=q) | Q(rif__icontains=q),
                activo=True
            ).order_by('nombre')[:10]
            
            data = [{'id': p.id, 'nombre': p.nombre, 'rif': p.rif} for p in proveedores]
            print(f"Encontrados {len(data)} proveedores")
            return JsonResponse(data, safe=False)
        else:
            print("Query muy corta")
            return JsonResponse([], safe=False)
    except Exception as e:
        print(f"Error en buscar_proveedores: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def buscar_clientes(request):
    """Buscar clientes para autocompletado"""
    q = request.GET.get('q', '')
    if len(q) >= 2:
        clientes = Cliente.objects.filter(
            Q(nombre__icontains=q) | Q(rif__icontains=q),
            activo=True
        )[:10]
        data = [{'id': c.id, 'nombre': c.nombre, 'rif': c.rif} for c in clientes]
        return JsonResponse(data, safe=False)
    return JsonResponse([], safe=False)

@login_required
def buscar_productos(request):
    """Buscar productos para autocompletado"""
    q = request.GET.get('q', '')
    tipo_factura = request.GET.get('tipo', 'venta')  # Por defecto es venta
    
    if len(q) >= 2:
        # Filtro base
        filtro = Q(nombre__icontains=q) | Q(codigo__icontains=q)
        filtro &= Q(activo=True)
        
        # Para facturas de venta, solo mostrar productos con stock
        if tipo_factura == 'venta':
            filtro &= Q(stock__gt=0)
        
        productos = Producto.objects.filter(filtro)[:10]
        
        # Para facturas de compra, usar costo como precio; para venta, usar precio de venta
        data = []
        for p in productos:
            if tipo_factura == 'compra':
                data.append({
                    'id': p.id, 
                    'nombre': p.nombre, 
                    'codigo': p.codigo, 
                    'precio': p.costo,  # Usar costo para compras
                    'precio_venta': p.precio,  # Incluir precio de venta para edición
                    'stock': p.stock, 
                    'iva': p.iva
                })
            else:
                data.append({
                    'id': p.id, 
                    'nombre': p.nombre, 
                    'codigo': p.codigo, 
                    'precio': p.precio,  # Usar precio de venta para ventas
                    'stock': p.stock, 
                    'iva': p.iva
                })
        
        return JsonResponse(data, safe=False)
    return JsonResponse([], safe=False)

@login_required
def proveedor_crear_ajax(request):
    """Crear proveedor via AJAX"""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            rif = request.POST.get('rif', '').strip()
            direccion = request.POST.get('direccion', '').strip()
            telefono = request.POST.get('telefono', '').strip()
            email = request.POST.get('email', '').strip()
            
            # Validaciones
            if not nombre:
                return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'})
            
            if not rif:
                return JsonResponse({'success': False, 'error': 'El RIF es obligatorio'})
            
            # Verificar si ya existe un proveedor con ese RIF
            if Proveedor.objects.filter(rif=rif).exists():
                return JsonResponse({'success': False, 'error': 'Ya existe un proveedor con ese RIF'})
            
            # Crear el proveedor
            proveedor = Proveedor.objects.create(
                nombre=nombre,
                rif=rif,
                direccion=direccion,
                telefono=telefono,
                email=email,
                activo=True
            )
            
            return JsonResponse({
                'success': True,
                'proveedor': {
                    'id': proveedor.id,
                    'nombre': proveedor.nombre,
                    'rif': proveedor.rif
                }
            })
            
        except Exception as e:
            print(f"Error creando proveedor: {e}")
            return JsonResponse({'success': False, 'error': 'Error interno del servidor'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
def cliente_crear_ajax(request):
    """Crear cliente via AJAX"""
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            rif = request.POST.get('rif', '').strip()
            direccion = request.POST.get('direccion', '').strip()
            telefono = request.POST.get('telefono', '').strip()
            email = request.POST.get('email', '').strip()
            
            # Validaciones
            if not nombre:
                return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'})
            
            if not rif:
                return JsonResponse({'success': False, 'error': 'El RIF es obligatorio'})
            
            # Verificar si ya existe un cliente con ese RIF
            if Cliente.objects.filter(rif=rif).exists():
                return JsonResponse({'success': False, 'error': 'Ya existe un cliente con ese RIF'})
            
            # Validar email si se proporciona
            if email and '@' not in email:
                return JsonResponse({'success': False, 'error': 'El email debe tener un formato válido'})
            
            # Si no se proporciona email, usar uno por defecto
            if not email:
                email = f'{rif.lower()}@cliente.local'
            
            # Crear el cliente
            cliente = Cliente.objects.create(
                nombre=nombre,
                rif=rif,
                direccion=direccion,
                telefono=telefono,
                email=email,
                activo=True
            )
            
            return JsonResponse({
                'success': True,
                'cliente': {
                    'id': cliente.id,
                    'nombre': cliente.nombre,
                    'rif': cliente.rif
                }
            })
            
        except Exception as e:
            print(f"Error creando cliente: {e}")
            return JsonResponse({'success': False, 'error': f'Error interno del servidor: {str(e)}'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})




# ============================================================================
# EXPORTACIÓN A EXCEL
# ============================================================================

def exportar_facturas_excel(request):
    """Exportar facturas a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener parámetros de filtro
    tipo = request.GET.get('tipo', '')
    q = request.GET.get('q', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    # Filtrar facturas
    facturas = Factura.objects.all()
    if tipo:
        facturas = facturas.filter(tipo=tipo)
    if q:
        facturas = facturas.filter(
            Q(numero__icontains=q) |
            Q(proveedor__nombre__icontains=q) |
            Q(cliente__nombre__icontains=q)
        )
    if desde:
        facturas = facturas.filter(fecha__gte=desde)
    if hasta:
        facturas = facturas.filter(fecha__lte=hasta)
    
    facturas = facturas.order_by('-fecha')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Fecha', 'Tipo', 'Número', 'Proveedor/Cliente', 
        'Estado', 'Subtotal 5%', 'Subtotal 10%', 'IVA', 'Total'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, factura in enumerate(facturas, 2):
        ws.cell(row=row, column=1, value=factura.id)
        ws.cell(row=row, column=2, value=factura.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=factura.get_tipo_display())
        ws.cell(row=row, column=4, value=factura.numero or f"#{factura.id}")
        
        if factura.tipo == 'compra':
            ws.cell(row=row, column=5, value=str(factura.proveedor))
        else:
            ws.cell(row=row, column=5, value=str(factura.cliente))
        
        ws.cell(row=row, column=6, value=factura.get_estado_display())
        ws.cell(row=row, column=7, value=factura.subtotal_5 or 0)
        ws.cell(row=row, column=8, value=factura.subtotal_10 or 0)
        ws.cell(row=row, column=9, value=factura.iva or 0)
        ws.cell(row=row, column=10, value=factura.total or 0)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

def exportar_productos_excel(request):
    """Exportar productos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener parámetros de filtro
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    
    # Filtrar productos
    productos = Producto.objects.all()
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | Q(codigo__icontains=q)
        )
    if estado:
        if estado == 'normal':
            productos = productos.filter(stock__gt=F('stock_minimo'))
        elif estado == 'minimo':
            productos = productos.filter(stock=F('stock_minimo'))
        elif estado == 'critico':
            productos = productos.filter(stock__lt=F('stock_minimo'))
    
    productos = productos.order_by('nombre')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'Código', 'Nombre', 'Stock', 'Stock Mínimo', 'Costo', 
        'Precio', 'IVA (%)', 'Estado'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.codigo)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=producto.stock)
        ws.cell(row=row, column=4, value=producto.stock_minimo)
        ws.cell(row=row, column=5, value=producto.costo or 0)
        ws.cell(row=row, column=6, value=producto.precio or 0)
        ws.cell(row=row, column=7, value=producto.iva)
        
        # Estado
        if producto.stock <= producto.stock_minimo:
            estado = "Stock Crítico" if producto.stock < producto.stock_minimo else "Stock Mínimo"
        else:
            estado = "Stock Normal"
        ws.cell(row=row, column=8, value=estado)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

def exportar_proveedores_excel(request):
    """Exportar proveedores a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener parámetros de filtro
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    
    # Filtrar proveedores
    proveedores = Proveedor.objects.all()
    if q:
        proveedores = proveedores.filter(
            Q(nombre__icontains=q) | Q(rif__icontains=q)
        )
    if estado:
        if estado == 'activo':
            proveedores = proveedores.filter(activo=True)
        elif estado == 'inactivo':
            proveedores = proveedores.filter(activo=False)
    
    proveedores = proveedores.order_by('nombre')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'Nombre', 'RIF/RUC', 'Dirección', 'Teléfono', 'Email', 
        'Estado', 'Saldo'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, proveedor in enumerate(proveedores, 2):
        ws.cell(row=row, column=1, value=proveedor.nombre)
        ws.cell(row=row, column=2, value=proveedor.rif)
        ws.cell(row=row, column=3, value=proveedor.direccion or '')
        ws.cell(row=row, column=4, value=proveedor.telefono or '')
        ws.cell(row=row, column=5, value=proveedor.email or '')
        ws.cell(row=row, column=6, value="Activo" if proveedor.activo else "Inactivo")
        ws.cell(row=row, column=7, value=proveedor.saldo or 0)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="proveedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

def exportar_detalles_facturas_excel(request):
    """Exportar detalles de facturas a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener parámetros de filtro
    q = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    # Filtrar detalles
    detalles = DetalleFactura.objects.all()
    if q:
        detalles = detalles.filter(
            Q(producto__nombre__icontains=q) | Q(producto__codigo__icontains=q)
        )
    if tipo:
        detalles = detalles.filter(factura__tipo=tipo)
    if desde:
        detalles = detalles.filter(factura__fecha__gte=desde)
    if hasta:
        detalles = detalles.filter(factura__fecha__lte=hasta)
    
    detalles = detalles.order_by('-factura__fecha')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalles Facturas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'Fecha Factura', 'Número Factura', 'Tipo', 'Producto', 'Código', 
        'Cantidad', 'Precio Unitario', 'Subtotal', 'IVA', 'Total'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, detalle in enumerate(detalles, 2):
        ws.cell(row=row, column=1, value=detalle.factura.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=detalle.factura.numero or f"#{detalle.factura.id}")
        ws.cell(row=row, column=3, value=detalle.factura.get_tipo_display())
        ws.cell(row=row, column=4, value=detalle.producto.nombre)
        ws.cell(row=row, column=5, value=detalle.producto.codigo)
        ws.cell(row=row, column=6, value=detalle.cantidad)
        ws.cell(row=row, column=7, value=detalle.precio_unitario)
        ws.cell(row=row, column=8, value=detalle.subtotal)
        ws.cell(row=row, column=9, value=detalle.iva)
        ws.cell(row=row, column=10, value=detalle.total)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="detalles_facturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

# ============================================================================
# SISTEMA DE ALERTAS
# ============================================================================

def obtener_alertas_stock(request):
    """Obtener alertas de stock bajo y productos agotados"""
    from django.db.models import F
    
    # Productos con stock bajo (stock <= stock_minimo)
    productos_stock_bajo = Producto.objects.filter(
        activo=True,
        stock__lte=F('stock_minimo'),
        stock__gt=0
    ).order_by('stock')
    
    # Productos agotados (stock = 0)
    productos_agotados = Producto.objects.filter(
        activo=True,
        stock=0
    ).order_by('nombre')
    
    # Facturas pendientes de pago (más de 30 días)
    from datetime import datetime, timedelta
    fecha_limite = datetime.now() - timedelta(days=30)
    facturas_vencidas = Factura.objects.filter(
        estado='pendiente',
        fecha__lt=fecha_limite
    ).order_by('fecha')
    
    alertas = {
        'stock_bajo': productos_stock_bajo,
        'agotados': productos_agotados,
        'facturas_vencidas': facturas_vencidas,
        'total_alertas': productos_stock_bajo.count() + productos_agotados.count() + facturas_vencidas.count()
    }
    
    return alertas

def crear_notificacion(mensaje, tipo='warning', usuario=None):
    """Crear una notificación en el sistema"""
    from .models import Notificacion
    
    notificacion = Notificacion.objects.create(
        mensaje=mensaje,
        tipo=tipo,
        usuario=usuario
    )
    return notificacion

def verificar_alertas_stock():
    """Verificar y crear alertas automáticas de stock"""
    from django.db.models import F
    from datetime import datetime
    
    # Verificar productos con stock bajo
    productos_stock_bajo = Producto.objects.filter(
        activo=True,
        stock__lte=F('stock_minimo'),
        stock__gt=0
    )
    
    for producto in productos_stock_bajo:
        mensaje = f"Stock bajo: {producto.nombre} (Código: {producto.codigo}) - Stock actual: {producto.stock}, Mínimo: {producto.stock_minimo}"
        crear_notificacion(mensaje, 'warning')
    
    # Verificar productos agotados
    productos_agotados = Producto.objects.filter(
        activo=True,
        stock=0
    )
    
    for producto in productos_agotados:
        mensaje = f"Producto agotado: {producto.nombre} (Código: {producto.codigo}) - Stock: 0"
        crear_notificacion(mensaje, 'error')
    
    # Verificar facturas vencidas
    from datetime import timedelta
    fecha_limite = datetime.now() - timedelta(days=30)
    facturas_vencidas = Factura.objects.filter(
        estado='pendiente',
        fecha__lt=fecha_limite
    )
    
    for factura in facturas_vencidas:
        dias_vencida = (datetime.now() - factura.fecha.replace(tzinfo=None)).days
        mensaje = f"Factura vencida: #{factura.id} - {factura.get_tipo_display()} - Vencida hace {dias_vencida} días - Total: Gs. {factura.total:,}"
        crear_notificacion(mensaje, 'error')

@login_required
def notificaciones_list(request):
    """Lista de notificaciones del sistema"""
    # Obtener todas las notificaciones (globales y del usuario)
    notificaciones = Notificacion.objects.filter(
        models.Q(usuario=request.user) | models.Q(usuario__isnull=True)
    ).order_by('-fecha')
    
    # Marcar como leídas solo las del usuario actual
    notificaciones.filter(usuario=request.user).update(leida=True)
    
    context = {
        'notificaciones': notificaciones,
        'titulo': 'Notificaciones'
    }
    
    return render(request, 'notificaciones_list.html', context)

@login_required
def marcar_notificacion_leida(request, pk):
    """Marcar una notificación como leída"""
    notificacion = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
    notificacion.leida = True
    notificacion.save()
    
    return JsonResponse({'success': True})

@login_required
def obtener_notificaciones_ajax(request):
    """Obtener notificaciones no leídas para AJAX"""
    notificaciones = Notificacion.objects.filter(
        leida=False
    ).order_by('-fecha')[:5]
    
    data = []
    for notif in notificaciones:
        data.append({
            'id': notif.id,
            'mensaje': notif.mensaje,
            'tipo': notif.tipo,
            'fecha': notif.fecha.strftime('%d/%m/%Y %H:%M')
        })
    
    return JsonResponse({'notificaciones': data, 'count': notificaciones.count()})

# ============================================================================
# SISTEMA DE EMAIL
# ============================================================================

from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from datetime import datetime, timedelta

def enviar_email_alertas(alertas, destinatarios=None):
    """Enviar email con alertas del sistema"""
    if not getattr(settings, 'EMAIL_NOTIFICATIONS_ENABLED', False):
        return False
    
    if not destinatarios:
        destinatarios = [settings.EMAIL_ADMIN_ADDRESS]
    
    # Calcular días vencidos para facturas
    for factura in alertas.get('facturas_vencidas', []):
        factura.dias_vencida = (datetime.now() - factura.fecha.replace(tzinfo=None)).days
    
    # Renderizar template HTML
    html_content = render_to_string('emails/alertas_diarias.html', {
        'alertas': alertas,
        'total_alertas': alertas['total_alertas'],
        'fecha': datetime.now().strftime('%d/%m/%Y %H:%M'),
    })
    
    # Crear versión texto plano
    text_content = strip_tags(html_content)
    
    # Configurar email
    subject = f"🚨 Alertas del Sistema Avícola CVA - {datetime.now().strftime('%d/%m/%Y')}"
    from_email = settings.EMAIL_FROM_ADDRESS
    
    # Enviar email
    try:
        msg = EmailMultiAlternatives(subject, text_content, from_email, destinatarios)
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        return True
    except Exception as e:
        print(f"Error enviando email: {e}")
        return False

def enviar_email_stock_bajo(producto):
    """Enviar email específico para stock bajo"""
    if not getattr(settings, 'EMAIL_NOTIFICATIONS_ENABLED', False):
        return False
    
    subject = f"⚠️ Stock Bajo: {producto.nombre}"
    html_content = render_to_string('emails/stock_bajo.html', {
        'producto': producto,
        'fecha': datetime.now().strftime('%d/%m/%Y %H:%M'),
    })
    text_content = strip_tags(html_content)
    
    try:
        msg = EmailMultiAlternatives(
            subject, 
            text_content, 
            settings.EMAIL_FROM_ADDRESS, 
            [settings.EMAIL_ADMIN_ADDRESS]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        return True
    except Exception as e:
        print(f"Error enviando email de stock bajo: {e}")
        return False

def enviar_email_producto_agotado(producto):
    """Enviar email específico para producto agotado"""
    if not getattr(settings, 'EMAIL_NOTIFICATIONS_ENABLED', False):
        return False
    
    subject = f"❌ Producto Agotado: {producto.nombre}"
    html_content = render_to_string('emails/producto_agotado.html', {
        'producto': producto,
        'fecha': datetime.now().strftime('%d/%m/%Y %H:%M'),
    })
    text_content = strip_tags(html_content)
    
    try:
        msg = EmailMultiAlternatives(
            subject, 
            text_content, 
            settings.EMAIL_FROM_ADDRESS, 
            [settings.EMAIL_ADMIN_ADDRESS]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        return True
    except Exception as e:
        print(f"Error enviando email de producto agotado: {e}")
        return False

def enviar_email_factura_vencida(factura):
    """Enviar email específico para factura vencida"""
    if not getattr(settings, 'EMAIL_NOTIFICATIONS_ENABLED', False):
        return False
    
    dias_vencida = (datetime.now() - factura.fecha.replace(tzinfo=None)).days
    subject = f"💰 Factura Vencida #{factura.numero} - {dias_vencida} días"
    
    html_content = render_to_string('emails/factura_vencida.html', {
        'factura': factura,
        'dias_vencida': dias_vencida,
        'fecha': datetime.now().strftime('%d/%m/%Y %H:%M'),
    })
    text_content = strip_tags(html_content)
    
    try:
        msg = EmailMultiAlternatives(
            subject, 
            text_content, 
            settings.EMAIL_FROM_ADDRESS, 
            [settings.EMAIL_ADMIN_ADDRESS]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        return True
    except Exception as e:
        print(f"Error enviando email de factura vencida: {e}")
        return False

@login_required
def enviar_alertas_email(request):
    """Vista para enviar alertas por email manualmente"""
    alertas = obtener_alertas_stock(request)
    
    if enviar_email_alertas(alertas):
        messages.success(request, 'Email de alertas enviado correctamente.')
    else:
        messages.error(request, 'Error al enviar el email de alertas.')
    
    return redirect('dashboard')

# ============================================================================
# PANEL DE CONFIGURACIÓN
# ============================================================================

@login_required
def configuracion_panel(request):
    """Panel principal de configuración del sistema"""
    from .models import ConfiguracionSistema
    
    # Obtener configuraciones por categoría
    configuraciones = {}
    for categoria, nombre in ConfiguracionSistema.CATEGORIA_CHOICES:
        configs = ConfiguracionSistema.objects.filter(categoria=categoria, activo=True)
        configuraciones[categoria] = {
            'nombre': nombre,
            'configs': configs
        }
    
    # Configuraciones por defecto si no existen
    configuraciones_por_defecto = {
        'general': [
            ('nombre_empresa', 'Avícola CVA', 'Nombre de la empresa'),
            ('moneda', 'Gs.', 'Símbolo de moneda'),
            ('pais', 'Paraguay', 'País de la empresa'),
        ],
        'alertas': [
            ('frecuencia_alertas', '24', 'Frecuencia de alertas en horas'),
            ('dias_factura_vencida', '30', 'Días para considerar factura vencida'),
            ('alertas_stock_bajo', 'true', 'Habilitar alertas de stock bajo'),
            ('alertas_productos_agotados', 'true', 'Habilitar alertas de productos agotados'),
            ('alertas_facturas_vencidas', 'true', 'Habilitar alertas de facturas vencidas'),
        ],
        'email': [
            ('email_notificaciones', 'true', 'Habilitar notificaciones por email'),
            ('email_frecuencia', '24', 'Frecuencia de emails en horas'),
            ('email_destinatarios', 'admin@avicolacva.com', 'Emails destinatarios (separados por comas)'),
        ],
        'stock': [
            ('stock_minimo_default', '10', 'Stock mínimo por defecto'),
            ('alertas_stock_critico', 'true', 'Alertas de stock crítico'),
            ('stock_critico_porcentaje', '20', 'Porcentaje para stock crítico'),
        ],
        'facturacion': [
            ('iva_default', '10', 'IVA por defecto (%)'),
            ('numero_factura_inicial', '1', 'Número inicial de facturas'),
            ('formato_factura', 'FAC-{numero}', 'Formato de número de factura'),
        ],
        'tema': [
            ('tema_visual', 'azul', 'Tema visual del sistema (azul, oscuro, minimalista)'),
        ]
    }
    
    # Crear configuraciones por defecto si no existen
    for categoria, configs in configuraciones_por_defecto.items():
        for clave, valor_por_defecto, descripcion in configs:
            ConfiguracionSistema.set_valor(
                clave=clave,
                valor=valor_por_defecto,
                descripcion=descripcion,
                categoria=categoria
            )
    
    # Recargar configuraciones después de crear las por defecto
    for categoria, nombre in ConfiguracionSistema.CATEGORIA_CHOICES:
        configs = ConfiguracionSistema.objects.filter(categoria=categoria, activo=True)
        configuraciones[categoria] = {
            'nombre': nombre,
            'configs': configs
        }
    
    context = {
        'configuraciones': configuraciones,
        'titulo': 'Configuración del Sistema'
    }
    
    return render(request, 'configuracion_panel.html', context)

@login_required
def configuracion_editar(request, pk):
    """Editar una configuración específica"""
    from .models import ConfiguracionSistema
    
    config = get_object_or_404(ConfiguracionSistema, pk=pk)
    
    if request.method == 'POST':
        valor = request.POST.get('valor')
        descripcion = request.POST.get('descripcion')
        
        config.valor = valor
        config.descripcion = descripcion
        config.save()
        
        messages.success(request, f'Configuración "{config.clave}" actualizada correctamente.')
        return redirect('configuracion_panel')
    
    context = {
        'config': config,
        'titulo': f'Editar Configuración: {config.clave}'
    }
    
    return render(request, 'configuracion_editar.html', context)

@login_required
def configuracion_guardar_ajax(request):
    """Guardar configuración via AJAX"""
    from .models import ConfiguracionSistema
    import json
    
    if request.method == 'POST':
        # Intentar obtener datos JSON primero
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                clave = data.get('clave')
                valor = data.get('valor')
                categoria = data.get('categoria', 'general')
                descripcion = data.get('descripcion', '')
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'error': 'Datos JSON inválidos'
                })
        else:
            # Datos POST tradicionales
            clave = request.POST.get('clave')
            valor = request.POST.get('valor')
            categoria = request.POST.get('categoria', 'general')
            descripcion = request.POST.get('descripcion', '')
        
        if not clave or valor is None:
            return JsonResponse({
                'success': False,
                'error': 'Clave y valor son requeridos'
            })
        
        try:
            config = ConfiguracionSistema.set_valor(clave, valor, categoria=categoria, descripcion=descripcion)
            return JsonResponse({
                'success': True,
                'message': f'Configuración "{clave}" guardada correctamente.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al guardar configuración: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
def configuracion_resetear(request):
    """Resetear configuraciones a valores por defecto"""
    from .models import ConfiguracionSistema
    
    # Configuraciones por defecto
    configuraciones_por_defecto = {
        'general': [
            ('nombre_empresa', 'Avícola CVA', 'Nombre de la empresa'),
            ('moneda', 'Gs.', 'Símbolo de moneda'),
            ('pais', 'Paraguay', 'País de la empresa'),
        ],
        'tema': [
            ('tema_visual', 'azul', 'Tema visual del sistema'),
        ],
        'alertas': [
            ('frecuencia_alertas', '24', 'Frecuencia de alertas en horas'),
            ('dias_factura_vencida', '30', 'Días para considerar factura vencida'),
            ('alertas_stock_bajo', 'true', 'Habilitar alertas de stock bajo'),
            ('alertas_productos_agotados', 'true', 'Habilitar alertas de productos agotados'),
            ('alertas_facturas_vencidas', 'true', 'Habilitar alertas de facturas vencidas'),
        ],
        'email': [
            ('email_notificaciones', 'true', 'Habilitar notificaciones por email'),
            ('email_frecuencia', '24', 'Frecuencia de emails en horas'),
            ('email_destinatarios', 'admin@avicolacva.com', 'Emails destinatarios (separados por comas)'),
        ],
        'stock': [
            ('stock_minimo_default', '10', 'Stock mínimo por defecto'),
            ('alertas_stock_critico', 'true', 'Alertas de stock crítico'),
            ('stock_critico_porcentaje', '20', 'Porcentaje para stock crítico'),
        ],
        'facturacion': [
            ('iva_default', '10', 'IVA por defecto (%)'),
            ('numero_factura_inicial', '1', 'Número inicial de facturas'),
            ('formato_factura', 'FAC-{numero}', 'Formato de número de factura'),
        ]
    }
    
    # Resetear todas las configuraciones
    for categoria, configs in configuraciones_por_defecto.items():
        for clave, valor_por_defecto, descripcion in configs:
            ConfiguracionSistema.set_valor(
                clave=clave,
                valor=valor_por_defecto,
                descripcion=descripcion,
                categoria=categoria
            )
    
    messages.success(request, 'Configuraciones reseteadas a valores por defecto.')
    return redirect('configuracion_panel')

@login_required
def configuracion_temas(request):
    """Configuración de temas visuales"""
    tema_actual = ConfiguracionSistema.get_valor('tema_visual', 'azul')
    
    context = {
        'tema_actual': tema_actual,
    }
    
    return render(request, 'configuracion_temas.html', context)

# ============================================================================
# REPORTES AVANZADOS
# ============================================================================

@login_required
def reportes_dashboard(request):
    """Dashboard principal de reportes"""
    
    # Estadísticas básicas
    total_ventas = Factura.objects.filter(tipo='venta', estado='pagada').aggregate(
        total=Sum('total')
    )['total'] or 0
    
    total_facturas = Factura.objects.filter(tipo='venta').count()
    productos_activos = Producto.objects.filter(stock__gt=0).count()
    valor_inventario = Producto.objects.aggregate(
        valor_total=Sum(F('stock') * F('precio'))
    )['valor_total'] or 0
    
    # Productos más vendidos
    productos_vendidos = DetalleFactura.objects.filter(
        factura__tipo='venta',
        factura__estado='pagada'
    ).values('producto__nombre').annotate(
        cantidad_vendida=Sum('cantidad')
    ).order_by('-cantidad_vendida')[:5]
    
    # Clientes con más compras
    clientes_compras = Factura.objects.filter(
        tipo='venta',
        estado='pagada'
    ).values('cliente__nombre').annotate(
        total_compras=Sum('total')
    ).order_by('-total_compras')[:5]
    
    context = {
        'total_ventas': total_ventas,
        'total_facturas': total_facturas,
        'productos_activos': productos_activos,
        'valor_inventario': valor_inventario,
        'productos_vendidos': productos_vendidos,
        'clientes_compras': clientes_compras,
        'titulo': 'Reportes Avanzados'
    }
    
    return render(request, 'reportes_dashboard.html', context)

@login_required
def reporte_ventas_detallado(request):
    """Reporte detallado de ventas"""
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_factura = request.GET.get('tipo', 'venta')
    
    # Construir query
    query = Factura.objects.filter(tipo=tipo_factura)
    
    if fecha_inicio:
        query = query.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        query = query.filter(fecha__lte=fecha_fin)
    
    # Obtener datos
    facturas = query.order_by('-fecha')
    total_facturas = facturas.count()
    total_ventas = facturas.aggregate(total=Sum('total'))['total'] or 0
    
    context = {
        'facturas': facturas,
        'total_facturas': total_facturas,
        'total_ventas': total_ventas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo_factura': tipo_factura,
        'titulo': 'Reporte Detallado de Ventas'
    }
    
    return render(request, 'reporte_ventas_detallado.html', context)

@login_required
def reporte_productos_analisis(request):
    """Análisis de productos"""
    # Productos más vendidos
    productos_mas_vendidos = DetalleFactura.objects.filter(
        factura__tipo='venta',
        factura__estado='pagada'
    ).values('producto__nombre').annotate(
        cantidad_vendida=Sum('cantidad'),
        total_vendido=Sum('total')
    ).order_by('-cantidad_vendida')[:20]
    
    # Productos con stock bajo
    productos_stock_bajo = Producto.objects.filter(
        stock__lte=F('stock_minimo')
    ).values('nombre', 'stock', 'stock_minimo', 'precio')
    
    # Productos agotados
    productos_agotados = Producto.objects.filter(stock=0).values('nombre', 'precio')
    
    context = {
        'productos_mas_vendidos': productos_mas_vendidos,
        'productos_stock_bajo': productos_stock_bajo,
        'productos_agotados': productos_agotados,
        'titulo': 'Análisis de Productos'
    }
    
    return render(request, 'reporte_productos_analisis.html', context)

@login_required
def reporte_clientes_proveedores(request):
    """Reporte de clientes y proveedores"""
    # Análisis de clientes
    clientes_analisis = Factura.objects.filter(
        tipo='venta',
        estado='pagada'
    ).values('cliente__nombre').annotate(
        total_compras=Sum('total'),
        cantidad_facturas=Count('id')
    ).order_by('-total_compras')
    
    # Análisis de proveedores
    proveedores_analisis = Factura.objects.filter(
        tipo='compra',
        estado='pagada'
    ).values('proveedor__nombre').annotate(
        total_ventas=Sum('total'),
        cantidad_facturas=Count('id')
    ).order_by('-total_ventas')
    
    # Facturas vencidas
    facturas_vencidas = Factura.objects.filter(
        estado='pendiente',
        fecha__lt=datetime.now() - timedelta(days=30)
    ).select_related('cliente', 'proveedor')
    
    context = {
        'clientes_analisis': clientes_analisis,
        'proveedores_analisis': proveedores_analisis,
        'facturas_vencidas': facturas_vencidas,
        'titulo': 'Análisis de Clientes y Proveedores'
    }
    
    return render(request, 'reporte_clientes_proveedores.html', context)

@login_required
def pago_multiple_crear(request):
    """Crear un pago que puede asignarse a múltiples facturas"""
    proveedor_id = request.GET.get('proveedor')
    cliente_id = request.GET.get('cliente')
    
    proveedor = None
    cliente = None
    
    if proveedor_id:
        proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
    elif cliente_id:
        cliente = get_object_or_404(Cliente, pk=cliente_id)
    else:
        messages.error(request, 'Debe seleccionar un proveedor o cliente.')
        return redirect('factura_list')
    
    if request.method == 'POST':
        form = PagoMultipleForm(request.POST, proveedor=proveedor, cliente=cliente)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.usuario = request.user
            pago.save()
            
            # Asignación automática desde la factura más antigua
            monto_disponible = pago.monto_total
            facturas_asignadas = []
            
            # Obtener facturas pendientes ordenadas por fecha (más antigua primero)
            if proveedor:
                facturas_pendientes = Factura.objects.filter(
                    tipo='compra', 
                    proveedor=proveedor, 
                    estado='pendiente'
                ).order_by('fecha')
            else:
                facturas_pendientes = Factura.objects.filter(
                    tipo='venta', 
                    cliente=cliente, 
                    estado='pendiente'
                ).order_by('fecha')
            
            # Asignar montos automáticamente
            for factura in facturas_pendientes:
                if monto_disponible <= 0:
                    break
                
                saldo_factura = factura.saldo_pendiente
                monto_a_asignar = min(monto_disponible, saldo_factura)
                
                try:
                    pago.asignar_a_factura(factura, monto_a_asignar)
                    facturas_asignadas.append({
                        'numero': factura.numero,
                        'monto': monto_a_asignar
                    })
                    monto_disponible -= monto_a_asignar
                except ValueError as e:
                    messages.warning(request, f'Error al asignar factura #{factura.numero}: {str(e)}')
                    continue
            
            # Mensaje de éxito con detalles
            if facturas_asignadas:
                detalles = ', '.join([f"#{f['numero']} (Gs. {f['monto']:,})" for f in facturas_asignadas])
                messages.success(request, f'Pago creado y asignado automáticamente a: {detalles}')
                
                if monto_disponible > 0:
                    messages.info(request, f'Queda disponible Gs. {monto_disponible:,} para asignar a más facturas.')
                    return redirect('pago_asignar_facturas', pago_id=pago.pk)
                else:
                    messages.success(request, 'Pago completamente asignado.')
                    return redirect('pago_ver', pago_id=pago.pk)
            else:
                messages.warning(request, 'Pago creado pero no se pudo asignar a ninguna factura.')
                return redirect('pago_asignar_facturas', pago_id=pago.pk)
    else:
        form = PagoMultipleForm(proveedor=proveedor, cliente=cliente)
    
    # Obtener facturas pendientes para mostrar información
    if proveedor:
        facturas_pendientes = Factura.objects.filter(
            tipo='compra', 
            proveedor=proveedor, 
            estado='pendiente'
        ).order_by('fecha')
        total_pendiente = sum(factura.saldo_pendiente for factura in facturas_pendientes)
    else:
        facturas_pendientes = Factura.objects.filter(
            tipo='venta', 
            cliente=cliente, 
            estado='pendiente'
        ).order_by('fecha')
        total_pendiente = sum(factura.saldo_pendiente for factura in facturas_pendientes)
    
    context = {
        'form': form,
        'proveedor': proveedor,
        'cliente': cliente,
        'facturas_pendientes': facturas_pendientes,
        'total_pendiente': total_pendiente,
    }
    
    return render(request, 'pago_multiple_crear.html', context)

@login_required
def pago_asignar_facturas(request, pago_id):
    """Asignar un pago a múltiples facturas"""
    pago = get_object_or_404(Pago, pk=pago_id)
    
    # Determinar si es proveedor o cliente
    proveedor = None
    cliente = None
    
    # Buscar la primera factura pendiente para determinar el tipo
    if pago.pagos_facturas.exists():
        primera_factura = pago.pagos_facturas.first().factura
        if primera_factura.tipo == 'compra':
            proveedor = primera_factura.proveedor
        else:
            cliente = primera_factura.cliente
    else:
        # Buscar facturas pendientes del proveedor o cliente
        facturas_pendientes = Factura.objects.filter(estado='pendiente')
        if facturas_pendientes.exists():
            primera_factura = facturas_pendientes.first()
            if primera_factura.tipo == 'compra':
                proveedor = primera_factura.proveedor
            else:
                cliente = primera_factura.cliente
    
    if request.method == 'POST':
        # Procesar asignaciones individuales
        factura_ids = request.POST.getlist('factura_id')
        montos = request.POST.getlist('monto')
        
        # Validar que el total no exceda el monto del pago
        total_asignado = sum(int(monto) for monto in montos if monto)
        if total_asignado > pago.monto_total:
            messages.error(request, f'El total asignado (Gs. {total_asignado:,}) excede el monto del pago (Gs. {pago.monto_total:,})')
        else:
            # Crear las asignaciones
            for factura_id, monto in zip(factura_ids, montos):
                if monto and int(monto) > 0:
                    factura = get_object_or_404(Factura, pk=factura_id)
                    try:
                        pago.asignar_a_factura(factura, int(monto))
                    except ValueError as e:
                        messages.error(request, str(e))
                        break
            else:
                messages.success(request, 'Pago asignado correctamente a las facturas.')
                return redirect('pago_ver', pago_id=pago.pk)
    
    # Obtener facturas pendientes
    if proveedor:
        facturas_pendientes = Factura.objects.filter(
            tipo='compra', 
            proveedor=proveedor, 
            estado='pendiente'
        ).order_by('fecha')
    else:
        facturas_pendientes = Factura.objects.filter(
            tipo='venta', 
            cliente=cliente, 
            estado='pendiente'
        ).order_by('fecha')
    
    # Calcular total pendiente
    total_pendiente = sum(factura.saldo_pendiente for factura in facturas_pendientes)
    
    context = {
        'pago': pago,
        'proveedor': proveedor,
        'cliente': cliente,
        'facturas_pendientes': facturas_pendientes,
        'total_pendiente': total_pendiente,
    }
    
    return render(request, 'pago_asignar_facturas.html', context)

@login_required
def pago_ver(request, pago_id):
    """Ver detalles de un pago y sus asignaciones"""
    pago = get_object_or_404(Pago, pk=pago_id)
    
    # Calcular porcentaje asignado
    porcentaje_asignado = 0
    if pago.monto_total > 0:
        porcentaje_asignado = (pago.monto_asignado / pago.monto_total) * 100
    
    # Obtener asignaciones con pagado anterior calculado
    asignaciones = []
    for pago_factura in pago.pagos_facturas.all():
        pagado_anterior = pago_factura.factura.total_pagado - pago_factura.monto
        asignaciones.append({
            'pago_factura': pago_factura,
            'pagado_anterior': pagado_anterior
        })
    
    context = {
        'pago': pago,
        'asignaciones': asignaciones,
        'porcentaje_asignado': porcentaje_asignado,
    }
    
    return render(request, 'pago_ver.html', context)

@login_required
def pago_eliminar(request, pk):
    """Eliminar un pago y restaurar saldos"""
    pago = get_object_or_404(Pago, pk=pk)
    
    if request.method == 'POST':
        # Restaurar saldos de todas las facturas afectadas
        for pago_factura in pago.pagos_facturas.all():
            factura = pago_factura.factura
            if factura.tipo == 'compra' and factura.proveedor:
                factura.proveedor.saldo += pago_factura.monto
                factura.proveedor.save()
            elif factura.tipo == 'venta' and factura.cliente:
                factura.cliente.saldo += pago_factura.monto
                factura.cliente.save()
            
            # Actualizar estado de la factura
            factura.actualizar_estado()
        
        # Eliminar el pago (esto también eliminará las asignaciones por CASCADE)
        pago.delete()
        
        messages.success(request, 'Pago eliminado correctamente.')
        return redirect('factura_list')
    
    return render(request, 'pago_confirm_delete.html', {'pago': pago})

@login_required
def asignacion_eliminar(request, pk):
    """Eliminar una asignación específica de pago a factura"""
    asignacion = get_object_or_404(PagoFactura, pk=pk)
    pago = asignacion.pago
    
    # Solo permitir eliminar asignaciones de facturas de compra (proveedores)
    if asignacion.factura.tipo != 'compra':
        messages.error(request, 'Solo se pueden eliminar asignaciones de facturas de compra (proveedores).')
        return redirect('pago_ver', pago_id=pago.pk)
    
    if request.method == 'POST':
        # Restaurar saldo del proveedor
        factura = asignacion.factura
        if factura.proveedor:
            factura.proveedor.saldo += asignacion.monto
            factura.proveedor.save()
        
        # Actualizar estado de la factura
        factura.actualizar_estado()
        
        # Eliminar la asignación
        asignacion.delete()
        
        messages.success(request, 'Asignación eliminada correctamente.')
        return redirect('pago_ver', pago_id=pago.pk)
    
    return render(request, 'asignacion_confirm_delete.html', {'asignacion': asignacion})

@login_required
def pagos_proveedores_dashboard(request):
    """Dashboard específico para pagos a proveedores"""
    from datetime import datetime, timedelta
    
    # Fechas para filtros
    hoy = datetime.now().date()
    inicio_mes = hoy.replace(day=1)
    
    # Métricas principales
    total_pagado_mes = Pago.objects.filter(
        fecha__gte=inicio_mes,
        proveedor__isnull=False
    ).aggregate(total=Sum('monto_total'))['total'] or 0
    
    total_pendiente = Proveedor.objects.filter(activo=True).aggregate(total=Sum('saldo'))['total'] or 0
    
    proveedores_pendientes = Proveedor.objects.filter(activo=True, saldo__gt=0).count()
    
    facturas_vencidas = Factura.objects.filter(
        tipo='compra',
        estado='pendiente',
        fecha_vencimiento__lt=hoy
    ).count()
    
    # Proveedores con saldo pendiente
    proveedores_con_saldo = Proveedor.objects.filter(
        activo=True, 
        saldo__gt=0
    ).annotate(
        facturas_pendientes_count=Count('facturas', filter=Q(facturas__estado='pendiente'))
    ).order_by('-saldo')[:10]
    
    # Facturas recientes
    facturas_recientes = Factura.objects.filter(
        tipo='compra'
    ).select_related('proveedor').order_by('-fecha')[:5]
    
    # Pagos recientes
    pagos_recientes = Pago.objects.filter(
        proveedor__isnull=False
    ).select_related('proveedor').order_by('-fecha')[:10]
    
    context = {
        'total_pagado': total_pagado_mes,
        'total_pendiente': total_pendiente,
        'proveedores_pendientes': proveedores_pendientes,
        'facturas_vencidas': facturas_vencidas,
        'proveedores_con_saldo': proveedores_con_saldo,
        'facturas_recientes': facturas_recientes,
        'pagos_recientes': pagos_recientes,
    }
    
    return render(request, 'pagos_proveedores_dashboard.html', context)

@login_required
def pago_proveedor_crear(request):
    """Crear pago específico para proveedores"""
    proveedor = None
    factura_especifica = None
    facturas_pendientes = []
    
    # Si se pasa un proveedor específico
    proveedor_id = request.GET.get('proveedor')
    if proveedor_id:
        proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
        facturas_pendientes = Factura.objects.filter(
            tipo='compra',
            proveedor=proveedor,
            estado='pendiente'
        ).order_by('fecha')
    
    # Si se pasa una factura específica
    factura_id = request.GET.get('factura')
    if factura_id:
        factura_especifica = get_object_or_404(Factura, pk=factura_id)
        # Solo permitir facturas de compra (proveedores)
        if factura_especifica.tipo != 'compra':
            messages.error(request, 'Solo se pueden realizar pagos a facturas de compra (proveedores).')
            return redirect('factura_ver', pk=factura_especifica.pk)
        proveedor = factura_especifica.proveedor
        # Obtener todas las facturas pendientes del proveedor, no solo la específica
        facturas_pendientes = Factura.objects.filter(
            tipo='compra',
            proveedor=proveedor,
            estado='pendiente'
        ).order_by('fecha')
    
    if request.method == 'POST':
        form = PagoMultipleForm(request.POST, proveedor=proveedor, factura_especifica=factura_especifica)
        if form.is_valid():
            pago = form.save(commit=False)
            pago.usuario = request.user
            
            # Establecer el proveedor en el pago
            if proveedor:
                pago.proveedor = proveedor
            else:
                messages.error(request, 'Error: No se especificó un proveedor para el pago.')
                return redirect('pagos_proveedores_dashboard')
            
            pago.save()
            
            # Asignación automática desde la factura más antigua
            monto_disponible = pago.monto_total
            facturas_asignadas = []
            
            # Si hay una factura específica, asignar solo a esa factura
            if factura_especifica:
                monto_a_asignar = min(monto_disponible, factura_especifica.saldo_pendiente)
                PagoFactura.objects.create(
                    pago=pago,
                    factura=factura_especifica,
                    monto=monto_a_asignar
                )
                facturas_asignadas.append({
                    'factura': factura_especifica.numero,
                    'monto': monto_a_asignar
                })
            else:
                # Asignar automáticamente a todas las facturas pendientes
                for factura in facturas_pendientes:
                    if monto_disponible <= 0:
                        break
                        
                    monto_a_asignar = min(monto_disponible, factura.saldo_pendiente)
                    
                    # Crear asignación
                    PagoFactura.objects.create(
                        pago=pago,
                        factura=factura,
                        monto=monto_a_asignar
                    )
                    
                    facturas_asignadas.append({
                        'factura': factura.numero,
                        'monto': monto_a_asignar
                    })
                    
                    monto_disponible -= monto_a_asignar
            
            # Actualizar saldo del proveedor
            proveedor.saldo = max(0, proveedor.saldo - pago.monto_total)
            proveedor.save()
            
            messages.success(
                request, 
                f'Pago creado exitosamente. Se asignaron {len(facturas_asignadas)} facturas automáticamente.'
            )
            
            # Redirigir a la página desde donde se accedió
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            elif factura_especifica:
                return redirect('factura_ver', pk=factura_especifica.pk)
            else:
                return redirect('pago_proveedor_ver', pago_id=pago.pk)
    else:
        form = PagoMultipleForm(proveedor=proveedor, factura_especifica=factura_especifica)
    
    context = {
        'form': form,
        'proveedor': proveedor,
        'factura_especifica': factura_especifica,
        'facturas_pendientes': facturas_pendientes,
    }
    
    return render(request, 'pago_proveedor_crear.html', context)



@login_required
def pago_proveedor_ver(request, pago_id):
    """Ver detalles de un pago específico a proveedor"""
    pago = get_object_or_404(Pago, pk=pago_id, proveedor__isnull=False)
    
    # Calcular porcentaje asignado
    porcentaje_asignado = 0
    if pago.monto_total > 0:
        porcentaje_asignado = (pago.monto_asignado / pago.monto_total) * 100
    
    # Obtener asignaciones con pagado anterior calculado
    asignaciones = []
    for pago_factura in pago.pagos_facturas.all():
        pagado_anterior = pago_factura.factura.total_pagado - pago_factura.monto
        asignaciones.append({
            'pago_factura': pago_factura,
            'pagado_anterior': pagado_anterior
        })
    
    context = {
        'pago': pago,
        'asignaciones': asignaciones,
        'porcentaje_asignado': porcentaje_asignado,
    }
    
    return render(request, 'pago_proveedor_ver.html', context)

@login_required
def pago_proveedor_asignar(request, pago_id):
    """Asignar facturas a un pago específico de proveedor"""
    pago = get_object_or_404(Pago, pk=pago_id, proveedor__isnull=False)
    
    if pago.monto_disponible <= 0:
        messages.warning(request, 'Este pago ya está completamente asignado.')
        next_url = request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('pago_proveedor_ver', pago_id=pago.pk)
    
    if request.method == 'POST':
        formset = PagoFacturaFormSet(request.POST, instance=pago)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Facturas asignadas correctamente.')
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('pago_proveedor_ver', pago_id=pago.pk)
    else:
        formset = PagoFacturaFormSet(instance=pago)
    
    # Obtener facturas pendientes del proveedor
    facturas_pendientes = Factura.objects.filter(
        tipo='compra',
        proveedor=pago.proveedor,
        estado='pendiente'
    ).exclude(
        pagos_facturas__pago=pago
    ).order_by('fecha')
    
    # Calcular total pendiente
    total_pendiente = sum(factura.saldo_pendiente for factura in facturas_pendientes)
    
    context = {
        'pago': pago,
        'formset': formset,
        'facturas_pendientes': facturas_pendientes,
        'total_pendiente': total_pendiente,
    }
    
    return render(request, 'pago_proveedor_asignar.html', context)

@login_required
def pago_proveedor_eliminar(request, pago_id):
    """Eliminar un pago específico de proveedor"""
    pago = get_object_or_404(Pago, pk=pago_id, proveedor__isnull=False)
    
    if request.method == 'POST':
        # Restaurar saldo del proveedor
        pago.proveedor.saldo += pago.monto_total
        pago.proveedor.save()
        
        # Eliminar el pago (las asignaciones se eliminan en cascada)
        pago.delete()
        
        messages.success(request, 'Pago eliminado correctamente.')
        next_url = request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('pagos_proveedores_dashboard')
    
    context = {
        'pago': pago,
    }
    
    return render(request, 'pago_proveedor_confirm_delete.html', context)

@login_required
def asignacion_proveedor_eliminar(request, asignacion_id):
    """Eliminar una asignación específica de pago a factura"""
    asignacion = get_object_or_404(PagoFactura, pk=asignacion_id)
    
    if request.method == 'POST':
        # Restaurar saldo del proveedor
        asignacion.pago.proveedor.saldo += asignacion.monto
        asignacion.pago.proveedor.save()
        
        # Eliminar la asignación
        asignacion.delete()
        
        messages.success(request, 'Asignación eliminada correctamente.')
        next_url = request.GET.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('pago_proveedor_ver', pago_id=asignacion.pago.pk)
    
    context = {
        'asignacion': asignacion,
    }
    
    return render(request, 'asignacion_proveedor_confirm_delete.html', context)

@login_required
def pagos_proveedores_reportes(request):
    """Reportes de pagos a proveedores"""
    from datetime import datetime, timedelta
    
    # Fechas para filtros
    hoy = datetime.now().date()
    inicio_mes = hoy.replace(day=1)
    inicio_anio = hoy.replace(month=1, day=1)
    
    # Estadísticas generales
    total_pagos = Pago.objects.filter(proveedor__isnull=False).count()
    total_monto_pagado = Pago.objects.filter(proveedor__isnull=False).aggregate(total=Sum('monto_total'))['total'] or 0
    
    # Pagos por mes (últimos 6 meses)
    pagos_por_mes = []
    meses_labels = []
    for i in range(6):
        fecha = hoy - timedelta(days=30*i)
        inicio_mes_grafico = fecha.replace(day=1)
        fin_mes_grafico = (inicio_mes_grafico + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        monto = Pago.objects.filter(
            proveedor__isnull=False,
            fecha__gte=inicio_mes_grafico,
            fecha__lte=fin_mes_grafico
        ).aggregate(total=Sum('monto_total'))['total'] or 0
        
        meses_labels.insert(0, fecha.strftime('%b %Y'))
        pagos_por_mes.insert(0, monto)
    
    # Top proveedores por monto pagado
    top_proveedores = Pago.objects.filter(
        proveedor__isnull=False
    ).values('proveedor__nombre').annotate(
        total_pagado=Sum('monto_total')
    ).order_by('-total_pagado')[:10]
    
    # Pagos por tipo
    pagos_por_tipo = Pago.objects.filter(
        proveedor__isnull=False
    ).values('tipo').annotate(
        total=Sum('monto_total'),
        count=Count('id')
    )
    
    context = {
        'total_pagos': total_pagos,
        'total_monto_pagado': total_monto_pagado,
        'meses_labels': meses_labels,
        'pagos_por_mes': pagos_por_mes,
        'top_proveedores': top_proveedores,
        'pagos_por_tipo': pagos_por_tipo,
    }
    
    return render(request, 'pagos_proveedores_reportes.html', context)

@login_required
def pagos_proveedores_vencidas(request):
    """Lista de facturas vencidas de proveedores"""
    from datetime import datetime
    
    hoy = datetime.now().date()
    
    # Facturas vencidas
    facturas_vencidas = Factura.objects.filter(
        tipo='compra',
        estado='pendiente',
        fecha_vencimiento__lt=hoy
    ).select_related('proveedor').order_by('fecha_vencimiento')
    
    # Agrupar por proveedor
    proveedores_vencidas = {}
    for factura in facturas_vencidas:
        if factura.proveedor not in proveedores_vencidas:
            proveedores_vencidas[factura.proveedor] = {
                'facturas': [],
                'total_vencido': 0,
                'dias_vencido_promedio': 0
            }
        
        dias_vencido = (hoy - factura.fecha_vencimiento).days
        proveedores_vencidas[factura.proveedor]['facturas'].append({
            'factura': factura,
            'dias_vencido': dias_vencido
        })
        proveedores_vencidas[factura.proveedor]['total_vencido'] += factura.saldo_pendiente
    
    # Calcular días vencido promedio por proveedor
    for proveedor_data in proveedores_vencidas.values():
        dias_totales = sum(f['dias_vencido'] for f in proveedor_data['facturas'])
        proveedor_data['dias_vencido_promedio'] = dias_totales / len(proveedor_data['facturas'])
    
    context = {
        'facturas_vencidas': facturas_vencidas,
        'proveedores_vencidas': proveedores_vencidas,
        'total_facturas_vencidas': facturas_vencidas.count(),
        'total_monto_vencido': sum(f.saldo_pendiente for f in facturas_vencidas),
    }
    
    return render(request, 'pagos_proveedores_vencidas.html', context)

@login_required
def reporte_pagos_proveedores(request):
    """Reporte avanzado de pagos a proveedores"""
    from datetime import datetime, timedelta
    from django.db.models import Q
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    proveedor_id = request.GET.get('proveedor')
    tipo_pago = request.GET.get('tipo_pago')
    
    # Fechas por defecto (último mes)
    hoy = datetime.now().date()
    if not fecha_inicio:
        fecha_inicio = (hoy - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_fin:
        fecha_fin = hoy.strftime('%Y-%m-%d')
    
    # Convertir fechas
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (hoy - timedelta(days=30)).date()
        fecha_fin_dt = hoy
    
    # Filtro base
    pagos_query = Pago.objects.filter(
        proveedor__isnull=False,
        fecha__date__gte=fecha_inicio_dt,
        fecha__date__lte=fecha_fin_dt
    ).select_related('proveedor', 'usuario')
    
    # Aplicar filtros adicionales
    if proveedor_id:
        pagos_query = pagos_query.filter(proveedor_id=proveedor_id)
    
    if tipo_pago:
        pagos_query = pagos_query.filter(tipo=tipo_pago)
    
    # Obtener pagos
    pagos = pagos_query.order_by('-fecha')
    
    # Estadísticas generales
    total_pagos = pagos.count()
    total_monto_pagado = pagos.aggregate(total=Sum('monto_total'))['total'] or 0
    total_asignado = pagos.aggregate(total=Sum('monto_asignado'))['total'] or 0
    total_disponible = total_monto_pagado - total_asignado
    
    # Estadísticas por proveedor
    stats_por_proveedor = pagos.values('proveedor__nombre').annotate(
        total_pagado=Sum('monto_total'),
        total_asignado=Sum('monto_asignado'),
        cantidad_pagos=Count('id')
    ).order_by('-total_pagado')
    
    # Estadísticas por tipo de pago
    stats_por_tipo = pagos.values('tipo').annotate(
        total=Sum('monto_total'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por mes
    stats_por_mes = pagos.extra(
        select={'mes': "DATE_TRUNC('month', fecha)"}
    ).values('mes').annotate(
        total=Sum('monto_total'),
        cantidad=Count('id')
    ).order_by('mes')
    
    # Top 10 pagos más grandes
    top_pagos = pagos.order_by('-monto_total')[:10]
    
    # Proveedores con más pagos
    proveedores_mas_pagos = pagos.values('proveedor__nombre').annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')[:10]
    
    # Facturas más pagadas
    facturas_mas_pagadas = PagoFactura.objects.filter(
        pago__in=pagos
    ).values('factura__numero', 'factura__proveedor__nombre').annotate(
        total_pagado=Sum('monto')
    ).order_by('-total_pagado')[:10]
    
    # Usuarios que realizaron pagos
    usuarios_pagos = pagos.values('usuario__username').annotate(
        total_pagado=Sum('monto_total'),
        cantidad=Count('id')
    ).order_by('-total_pagado')
    
    # Filtros disponibles
    proveedores_disponibles = Proveedor.objects.filter(activo=True).order_by('nombre')
    tipos_pago_disponibles = Pago.TIPO_CHOICES
    
    context = {
        'pagos': pagos,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'proveedor_seleccionado': proveedor_id,
        'tipo_pago_seleccionado': tipo_pago,
        
        # Estadísticas
        'total_pagos': total_pagos,
        'total_monto_pagado': total_monto_pagado,
        'total_asignado': total_asignado,
        'total_disponible': total_disponible,
        
        # Datos para gráficos y tablas
        'stats_por_proveedor': stats_por_proveedor,
        'stats_por_tipo': stats_por_tipo,
        'stats_por_mes': stats_por_mes,
        'top_pagos': top_pagos,
        'proveedores_mas_pagos': proveedores_mas_pagos,
        'facturas_mas_pagadas': facturas_mas_pagadas,
        'usuarios_pagos': usuarios_pagos,
        
        # Filtros
        'proveedores_disponibles': proveedores_disponibles,
        'tipos_pago_disponibles': tipos_pago_disponibles,
    }
    
    return render(request, 'reporte_pagos_proveedores.html', context)


# ============================================================================
# VISTAS DE CONTROL DE CAJA
# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================

@login_required
def caja_list(request):
    """Lista de cajas"""
    cajas = Caja.objects.select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Caja actual (hoy)
    from datetime import date
    caja_actual = cajas.filter(fecha=date.today()).first()
    
    context = {
        'cajas': cajas,
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'caja_actual': caja_actual,
    }
    
    return render(request, 'caja_list.html', context)


@login_required
def caja_abrir(request):
    """Abrir caja del día con denominaciones"""
    from datetime import date
    
    # Verificar si ya existe una caja para hoy
    caja_existente = Caja.objects.filter(fecha=date.today()).first()
    if caja_existente:
        messages.warning(request, 'Ya existe una caja abierta para hoy.')
        return redirect('caja_ver', caja_id=caja_existente.id)
    
    if request.method == 'POST':
        try:
            # Crear la caja
            caja = Caja.objects.create(
                fecha=date.today(),
                saldo_inicial=0,  # Se calculará automáticamente
                usuario_apertura=request.user
            )
            
            # Procesar denominaciones
            denominaciones_creadas = []
            total_calculado = 0
            
            for valor in Denominacion.VALOR_CHOICES:
                cantidad = request.POST.get(f'denominacion_{valor[0]}', 0)
                try:
                    cantidad = int(cantidad)
                    if cantidad > 0:
                        denominacion = Denominacion.objects.create(
                            caja=caja,
                            valor=valor[0],
                            cantidad=cantidad
                        )
                        denominaciones_creadas.append(denominacion)
                        total_calculado += valor[0] * cantidad
                except ValueError:
                    continue
            
            # Calcular saldo inicial basado en denominaciones
            caja.calcular_saldo_inicial_denominaciones()
            caja.save()
            
            messages.success(request, f'Caja abierta con saldo inicial de Gs. {caja.saldo_inicial:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except Exception as e:
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Preparar denominaciones para el formulario
    denominaciones = []
    for valor, label in Denominacion.VALOR_CHOICES:
        denominaciones.append({
            'valor': valor,
            'label': label,
            'cantidad': 0
        })
    
    context = {
        'fecha_actual': date.today(),
        'denominaciones': denominaciones,
    }
    return render(request, 'caja_abrir.html', context)


@login_required
def caja_ver(request, caja_id):
    """Ver detalles de una caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = caja.movimientos.select_related('usuario').order_by('-fecha')
    
    # Obtener gastos
    gastos = caja.gastos.select_related('usuario').order_by('-fecha')
    
    # Obtener denominaciones
    denominaciones = caja.denominaciones.all().order_by('-valor')
    
    # Calcular totales
    total_ingresos = movimientos.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
    total_egresos = movimientos.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or 0
    total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
    
    # Agrupar movimientos por categoría
    movimientos_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    context = {
        'caja': caja,
        'movimientos': movimientos,
        'gastos': gastos,
        'denominaciones': denominaciones,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_gastos': total_gastos,
        'movimientos_por_categoria': movimientos_por_categoria,
    }
    
    return render(request, 'caja_ver.html', context)


@login_required
def caja_cerrar(request, caja_id):
    """Cerrar caja con arqueo"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.warning(request, 'Esta caja ya está cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        saldo_real = request.POST.get('saldo_real', 0)
        observaciones = request.POST.get('observaciones', '')
        
        try:
            saldo_real = int(saldo_real)
            caja.cerrar_caja(saldo_real, request.user, observaciones)
            messages.success(request, 'Caja cerrada exitosamente.')
            return redirect('caja_ver', caja_id=caja.id)
        except ValueError:
            messages.error(request, 'El saldo real debe ser un número válido.')
    
    # Calcular saldo final esperado
    caja.calcular_saldo_final()
    
    context = {
        'caja': caja,
    }
    
    return render(request, 'caja_cerrar.html', context)


@login_required
def gasto_crear(request, caja_id):
    """Crear un nuevo gasto"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar gastos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto = Gasto.objects.create(
                caja=caja,
                categoria=categoria,
                descripcion=descripcion,
                monto=monto,
                comprobante=comprobante,
                observacion=observacion,
                usuario=request.user
            )
            
            messages.success(request, f'Gasto registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_crear.html', context)


@login_required
def gasto_editar(request, gasto_id):
    """Editar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden editar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        categoria = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        monto = request.POST.get('monto')
        comprobante = request.POST.get('comprobante', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            gasto.categoria = categoria
            gasto.descripcion = descripcion
            gasto.monto = monto
            gasto.comprobante = comprobante
            gasto.observacion = observacion
            gasto.save()
            
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('caja_ver', caja_id=gasto.caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'gasto': gasto,
        'categorias': Gasto.CATEGORIA_CHOICES,
    }
    
    return render(request, 'gasto_editar.html', context)


@login_required
def gasto_eliminar(request, gasto_id):
    """Eliminar un gasto"""
    gasto = get_object_or_404(Gasto, id=gasto_id)
    
    if gasto.caja.cerrada:
        messages.error(request, 'No se pueden eliminar gastos de una caja cerrada.')
        return redirect('caja_ver', caja_id=gasto.caja.id)
    
    if request.method == 'POST':
        caja_id = gasto.caja.id
        gasto.delete()
        messages.success(request, 'Gasto eliminado exitosamente.')
        return redirect('caja_ver', caja_id=caja_id)
    
    context = {
        'gasto': gasto,
    }
    
    return render(request, 'gasto_eliminar.html', context)


@login_required
def movimiento_crear(request, caja_id):
    """Crear un movimiento manual de caja"""
    caja = get_object_or_404(Caja, id=caja_id)
    
    if caja.cerrada:
        messages.error(request, 'No se pueden agregar movimientos a una caja cerrada.')
        return redirect('caja_ver', caja_id=caja.id)
    
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion')
        referencia = request.POST.get('referencia', '')
        observacion = request.POST.get('observacion', '')
        
        try:
            monto = int(monto)
            if monto <= 0:
                raise ValueError("El monto debe ser mayor a 0")
            
            movimiento = MovimientoCaja.registrar_movimiento(
                caja=caja,
                tipo=tipo,
                categoria=categoria,
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                referencia=referencia,
                observacion=observacion
            )
            
            messages.success(request, f'Movimiento registrado: {descripcion} - Gs. {monto:,}')
            return redirect('caja_ver', caja_id=caja.id)
            
        except ValueError as e:
            messages.error(request, f'Error: {str(e)}')
    
    context = {
        'caja': caja,
        'tipos': MovimientoCaja.TIPO_CHOICES,
        'categorias': MovimientoCaja.CATEGORIA_CHOICES,
    }
    
    return render(request, 'movimiento_crear.html', context)


@login_required
def reporte_caja(request):
    """Reporte de caja por período"""
    from datetime import datetime, timedelta
    
    # Parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    try:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    except ValueError:
        fecha_inicio_dt = (datetime.now() - timedelta(days=30)).date()
        fecha_fin_dt = datetime.now().date()
    
    # Obtener cajas del período
    cajas = Caja.objects.filter(
        fecha__gte=fecha_inicio_dt,
        fecha__lte=fecha_fin_dt
    ).select_related('usuario_apertura', 'usuario_cierre').order_by('-fecha')
    
    # Estadísticas generales
    total_cajas = cajas.count()
    cajas_abiertas = cajas.filter(cerrada=False).count()
    cajas_cerradas = cajas.filter(cerrada=True).count()
    
    # Totales financieros
    total_saldo_inicial = cajas.aggregate(total=Sum('saldo_inicial'))['total'] or 0
    total_saldo_final = cajas.aggregate(total=Sum('saldo_final'))['total'] or 0
    total_saldo_real = cajas.aggregate(total=Sum('saldo_real'))['total'] or 0
    total_diferencia = cajas.aggregate(total=Sum('diferencia'))['total'] or 0
    
    # Obtener todos los movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario')
    
    # Estadísticas por categoría
    stats_por_categoria = movimientos.values('categoria').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Estadísticas por tipo
    stats_por_tipo = movimientos.values('tipo').annotate(
        total=Sum('monto'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Top gastos
    top_gastos = Gasto.objects.filter(
        caja__fecha__gte=fecha_inicio_dt,
        caja__fecha__lte=fecha_fin_dt
    ).select_related('caja', 'usuario').order_by('-monto')[:10]
    
    context = {
        'cajas': cajas,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        
        # Estadísticas
        'total_cajas': total_cajas,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_saldo_inicial': total_saldo_inicial,
        'total_saldo_final': total_saldo_final,
        'total_saldo_real': total_saldo_real,
        'total_diferencia': total_diferencia,
        
        # Datos para análisis
        'stats_por_categoria': stats_por_categoria,
        'stats_por_tipo': stats_por_tipo,
        'top_gastos': top_gastos,
    }
    
    return render(request, 'reporte_caja.html', context)




# ============================================================================
# PAGOS
# ============================================================================

@login_required
def pagos_dashboard(request):
    """Dashboard principal del módulo de pagos"""
    # Obtener estadísticas de pagos
    total_pagos_clientes = Pago.objects.filter(factura__tipo='venta').count()
    total_pagos_proveedores = Pago.objects.filter(factura__tipo='compra').count()
    
    # Pagos recientes
    pagos_recientes = Pago.objects.select_related('factura', 'usuario').order_by('-fecha')[:10]
    
    # Facturas pendientes
    facturas_pendientes_clientes = Factura.objects.filter(
        tipo='venta', 
        estado='pendiente'
    ).select_related('cliente').order_by('-fecha')[:5]
    
    facturas_pendientes_proveedores = Factura.objects.filter(
        tipo='compra', 
        estado='pendiente'
    ).select_related('proveedor').order_by('-fecha')[:5]
    
    # Caja activa
    caja_activa = Caja.obtener_caja_activa()
    
    context = {
        'total_pagos_clientes': total_pagos_clientes,
        'total_pagos_proveedores': total_pagos_proveedores,
        'pagos_recientes': pagos_recientes,
        'facturas_pendientes_clientes': facturas_pendientes_clientes,
        'facturas_pendientes_proveedores': facturas_pendientes_proveedores,
        'caja_activa': caja_activa,
    }
    
    return render(request, 'pagos_dashboard.html', context)

@login_required
def pagos_clientes_list(request):
    """Lista de pagos de clientes (facturas de venta)"""
    # Obtener facturas de venta con pagos
    facturas_venta = Factura.objects.filter(
        tipo='venta'
    ).select_related('cliente').prefetch_related('pagos').order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        facturas_venta = facturas_venta.filter(
            Q(cliente__nombre__icontains=q) | 
            Q(numero__icontains=q) |
            Q(cliente__ruc__icontains=q)
        )
    
    if estado:
        facturas_venta = facturas_venta.filter(estado=estado)
    
    if desde:
        facturas_venta = facturas_venta.filter(fecha__date__gte=desde)
    
    if hasta:
        facturas_venta = facturas_venta.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(facturas_venta, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'facturas': page_obj,
        'tipo': 'venta',
        'titulo': 'Pagos de Clientes',
        'subtitulo': 'Facturas de Venta - Gestión de Cobros',
    }
    
    return render(request, 'pagos_list.html', context)

@login_required
def pagos_proveedores_list(request):
    """Lista de pagos a proveedores (facturas de compra)"""
    # Obtener facturas de compra con pagos
    facturas_compra = Factura.objects.filter(
        tipo='compra'
    ).select_related('proveedor').prefetch_related('pagos').order_by('-fecha')
    
    # Filtros
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')
    
    if q:
        facturas_compra = facturas_compra.filter(
            Q(proveedor__nombre__icontains=q) | 
            Q(numero__icontains=q) |
            Q(proveedor__ruc__icontains=q)
        )
    
    if estado:
        facturas_compra = facturas_compra.filter(estado=estado)
    
    if desde:
        facturas_compra = facturas_compra.filter(fecha__date__gte=desde)
    
    if hasta:
        facturas_compra = facturas_compra.filter(fecha__date__lte=hasta)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(facturas_compra, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'facturas': page_obj,
        'tipo': 'compra',
        'titulo': 'Pagos a Proveedores',
        'subtitulo': 'Facturas de Compra - Gestión de Pagos',
    }
    
    return render(request, 'pagos_list.html', context)



